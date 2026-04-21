import re
from typing import Any, Optional

from .prof_ratings import load_prof_ratings

def seed_offerings() -> list[dict[str, Any]]:
    """
    Demo-only catalog for gallery walk.
    Real product would ingest official schedule + catalog + prereq graph.
    """
    xlsx = _load_offerings_from_xlsx()
    if xlsx:
        return xlsx
    return [
        {
            "code": "COEN 11",
            "title": "Computer Programming I",
            "term": "2026 Spring",
            "units": 4,
            "prereqs": [],
            "schedule": "MWF 09:00-09:50",
            "quality": 4.2,
            "workload": 3.6,
        },
        {
            "code": "COEN 12",
            "title": "Computer Programming II",
            "term": "2026 Spring",
            "units": 4,
            "prereqs": ["COEN 11"],
            "schedule": "MWF 10:00-10:50",
            "quality": 4.0,
            "workload": 4.0,
        },
        {
            "code": "COEN 19",
            "title": "Mathematics for Computer Science",
            "term": "2026 Spring",
            "units": 4,
            "prereqs": [],
            "schedule": "TTh 11:00-12:15",
            "quality": 3.7,
            "workload": 4.2,
        },
        {
            "code": "COEN 20",
            "title": "Embedded Systems",
            "term": "2026 Spring",
            "units": 4,
            "prereqs": ["COEN 12"],
            "schedule": "TTh 13:00-14:15",
            "quality": 3.9,
            "workload": 4.4,
        },
        {
            "code": "COEN 21",
            "title": "Data Structures",
            "term": "2026 Spring",
            "units": 4,
            "prereqs": ["COEN 12"],
            "schedule": "MWF 13:00-13:50",
            "quality": 4.3,
            "workload": 4.6,
        },
        {
            "code": "COEN 146",
            "title": "Computer Networks",
            "term": "2026 Spring",
            "units": 4,
            "prereqs": ["COEN 12"],
            "schedule": "TTh 15:00-16:15",
            "quality": 4.1,
            "workload": 4.1,
        },
        {
            "code": "COEN 171",
            "title": "Software Engineering",
            "term": "2026 Spring",
            "units": 4,
            "prereqs": ["COEN 12"],
            "schedule": "MWF 15:00-15:50",
            "quality": 4.0,
            "workload": 3.9,
        },
        {
            "code": "MATH 53",
            "title": "Calculus IV",
            "term": "2026 Spring",
            "units": 4,
            "prereqs": ["MATH 52"],
            "schedule": "MWF 11:00-11:50",
            "quality": 3.6,
            "workload": 4.3,
        },
        {
            "code": "MATH 52",
            "title": "Calculus III",
            "term": "2026 Spring",
            "units": 4,
            "prereqs": ["MATH 51"],
            "schedule": "TTh 10:30-11:45",
            "quality": 3.5,
            "workload": 4.1,
        },
        {
            "code": "MATH 51",
            "title": "Calculus I",
            "term": "2026 Spring",
            "units": 4,
            "prereqs": [],
            "schedule": "TTh 09:00-10:15",
            "quality": 3.5,
            "workload": 4.0,
        },
        {
            "code": "CTW 1",
            "title": "Critical Thinking & Writing I",
            "term": "2026 Spring",
            "units": 4,
            "prereqs": [],
            "schedule": "MWF 14:00-14:50",
            "quality": 4.0,
            "workload": 3.2,
        },
        {
            "code": "CTW 2",
            "title": "Critical Thinking & Writing II",
            "term": "2026 Spring",
            "units": 4,
            "prereqs": ["CTW 1"],
            "schedule": "TTh 14:00-15:15",
            "quality": 3.9,
            "workload": 3.4,
        },
    ]


_CACHED_XLSX: Optional[list[dict[str, Any]]] = None


def _load_offerings_from_xlsx() -> Optional[list[dict[str, Any]]]:
    """
    If `SCU_Find_Course_Sections.xlsx` exists next to this prototype, load it as the
    course offering source (more realistic than hard-coded demo data).
    """
    global _CACHED_XLSX
    if _CACHED_XLSX is not None:
        return _CACHED_XLSX

    try:
        from pathlib import Path

        from openpyxl import load_workbook
    except Exception:
        _CACHED_XLSX = None
        return None

    try:
        base_dir = Path(__file__).resolve().parent.parent
        path = base_dir / "SCU_Find_Course_Sections.xlsx"
        if not path.exists():
            _CACHED_XLSX = None
            return None

        wb = load_workbook(str(path), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        prof_ratings = load_prof_ratings(base_dir)

        # Read header row using iter_rows (fast for large sheets)
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            _CACHED_XLSX = None
            return None

        header: dict[str, int] = {}
        for idx, v in enumerate(header_row):
            if isinstance(v, str) and v.strip():
                header[v.strip()] = idx  # 0-based

        def col(name: str) -> Optional[int]:
            return header.get(name)

        c_course_section = col("Course Section")
        c_course_subject = col("Course Subject")
        c_course_number = col("Course Number")
        c_units = col("Units")
        c_meeting = col("Meeting Patterns")
        c_tags = col("Course Tags")
        c_instructors = col("All Instructors")
        c_status = col("Section Status")
        c_enroll = col("Enrolled/Capacity")
        if c_course_section is None or c_course_subject is None or c_course_number is None:
            _CACHED_XLSX = None
            return None

        offerings_by_code: dict[str, dict[str, Any]] = {}
        instructors_by_code: dict[str, list[dict[str, Any]]] = {}

        # Example `Course Section`:
        # "ACTG 11-1 - Introduction to Financial Accounting (-)"
        sec_re = re.compile(r"^\s*([A-Z]{2,6})\s+(\d{1,4}[A-Z]{0,2})[-–]\d+\s*-\s*(.+?)\s*$")

        # Iterate values-only to avoid cell object overhead.
        for row_vals in ws.iter_rows(min_row=2, values_only=True):
            if not row_vals or c_course_section >= len(row_vals):
                continue
            sec = row_vals[c_course_section]
            if not isinstance(sec, str):
                continue
            m = sec_re.match(sec.strip())
            if not m:
                # Fallback: build code from "Course Number" + best-effort subject abbrev in `Course Section`
                m2 = re.match(r"^\s*([A-Z]{2,6})\s+(\d{1,4}[A-Z]{0,2})", sec.strip())
                if not m2:
                    continue
                dept = m2.group(1)
                num = m2.group(2)
                title = sec.split(" - ", 1)[1].strip() if " - " in sec else sec.strip()
            else:
                dept, num, title = m.group(1), m.group(2), m.group(3)

            code = f"{dept} {num}".upper()

            units_v = row_vals[c_units] if (c_units is not None and c_units < len(row_vals)) else None
            try:
                units = int(float(units_v)) if units_v is not None else 0
            except Exception:
                units = 0

            meeting = row_vals[c_meeting] if (c_meeting is not None and c_meeting < len(row_vals)) else ""
            tags = row_vals[c_tags] if (c_tags is not None and c_tags < len(row_vals)) else None
            tags_s = str(tags) if tags is not None else ""
            subj = (
                row_vals[c_course_subject]
                if (c_course_subject is not None and c_course_subject < len(row_vals))
                else ""
            )

            instr = (
                row_vals[c_instructors]
                if (c_instructors is not None and c_instructors < len(row_vals))
                else None
            )
            status = (
                row_vals[c_status] if (c_status is not None and c_status < len(row_vals)) else None
            )
            enroll = (
                row_vals[c_enroll] if (c_enroll is not None and c_enroll < len(row_vals)) else None
            )

            # Track instructors per section for per-course ranking.
            instr_s = str(instr).strip() if instr is not None else ""
            status_s = str(status).strip() if status is not None else ""
            enroll_s = str(enroll).strip() if enroll is not None else ""
            instructors_by_code.setdefault(code, []).append(
                {
                    "name": instr_s,
                    "status": status_s,
                    "enroll": enroll_s,
                    "meeting": str(meeting or ""),
                }
            )

            if code not in offerings_by_code:
                offerings_by_code[code] = {
                    "code": code,
                    "title": str(title).strip(),
                    # This export doesn't encode term in the sheet; assume this file is for current demo term.
                    "term": "2026 Spring",
                    "units": units or 4,
                    "prereqs": [],
                    "schedule": str(meeting or ""),
                    # Demo scoring signals (kept stable if not provided)
                    "quality": 4.0,
                    "workload": 3.8,
                    "tags": tags_s,
                    "subject": str(subj or "").strip(),
                }

        def enroll_ratio(s: str) -> float:
            # "28/25" => 1.12
            try:
                left, right = s.split("/", 1)
                a = float(left.strip())
                b = float(right.strip())
                return a / b if b else 999.0
            except Exception:
                return 999.0

        def status_rank(s: str) -> int:
            sl = (s or "").strip().lower()
            if sl == "open":
                return 0
            if sl == "waitlist":
                return 1
            if sl == "closed":
                return 2
            return 3

        def prof_rating(name: str) -> float:
            # unknown => -1 (sort last)
            key = " ".join(name.strip().lower().split())
            pr = prof_ratings.get(key)
            return float(pr.rating) if pr else -1.0

        # Attach ranked instructor list to each offering.
        offerings: list[dict[str, Any]] = []
        for code, o in offerings_by_code.items():
            sections = instructors_by_code.get(code, [])
            # De-dup instructor names while keeping best section (open + lower ratio)
            best_by_name: dict[str, dict[str, Any]] = {}
            for sec in sections:
                name = (sec.get("name") or "").strip()
                if not name:
                    continue
                cur = best_by_name.get(name)
                if cur is None:
                    best_by_name[name] = sec
                    continue
                if (status_rank(sec.get("status", "")), enroll_ratio(sec.get("enroll", ""))) < (
                    status_rank(cur.get("status", "")),
                    enroll_ratio(cur.get("enroll", "")),
                ):
                    best_by_name[name] = sec

            ranked = sorted(
                best_by_name.values(),
                # Higher professor rating first; then open-ness; then easier-to-get seats.
                key=lambda x: (
                    -prof_rating(str(x.get("name") or "")),
                    status_rank(x.get("status", "")),
                    enroll_ratio(x.get("enroll", "")),
                ),
            )
            o["instructors"] = [
                (
                    f"{x.get('name')}"
                    f"{' · ' + str(prof_rating(str(x.get('name') or ''))) if prof_rating(str(x.get('name') or '')) >= 0 else ''}"
                    f"{' (' + x.get('status') + ')' if x.get('status') else ''}"
                )
                for x in ranked[:6]
                if x.get("name")
            ]
            offerings.append(o)

        _CACHED_XLSX = offerings
        return offerings
    except Exception:
        _CACHED_XLSX = None
        return None
