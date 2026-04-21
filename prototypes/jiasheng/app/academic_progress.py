import re
from dataclasses import dataclass
from pathlib import Path
from typing import Optional, Union

from openpyxl import load_workbook


_COURSE_CODE_RE = re.compile(r"\b([A-Z]{2,6})\s+(\d{1,4}[A-Z]{0,2})\b")


@dataclass(frozen=True)
class AcademicProgress:
    completed_codes: set[str]
    unsatisfied_requirements: list[str]


def _norm_code(code: str) -> str:
    return re.sub(r"\s+", " ", code.strip().upper())


def load_academic_progress_xlsx(path: Union[str, Path]) -> AcademicProgress:
    """
    Parse SCU Workday export "View My Academic Progress" XLSX.

    We only need:
    - completed course codes (including In Progress, to avoid recommending duplicates)
    - unsatisfied requirement names (for tag matching / explanation)
    """
    p = Path(path)
    wb = load_workbook(str(p), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Find header row (usually row 2): contains "Requirement" and "Status"
    header_idx = None
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if not row:
            continue
        row_s = [str(x).strip() for x in row if x is not None]
        if "Requirement" in row_s and "Status" in row_s:
            header_idx = i
            header_row = list(row)
            break

    if header_idx is None or header_row is None:
        return AcademicProgress(completed_codes=set(), unsatisfied_requirements=[])

    header: dict[str, int] = {}
    for idx, v in enumerate(header_row):
        if isinstance(v, str) and v.strip():
            header[v.strip()] = idx

    def col(name: str) -> Optional[int]:
        return header.get(name)

    c_req = col("Requirement")
    c_status = col("Status")
    c_regs = col("Registrations Used")
    if c_req is None or c_status is None:
        return AcademicProgress(completed_codes=set(), unsatisfied_requirements=[])

    completed: set[str] = set()
    unsat: list[str] = []
    seen_unsat: set[str] = set()

    for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
        if not row:
            continue
        req = row[c_req] if c_req < len(row) else None
        status = row[c_status] if c_status < len(row) else None
        regs = row[c_regs] if (c_regs is not None and c_regs < len(row)) else None

        req_s = str(req).strip() if isinstance(req, str) else (str(req).strip() if req is not None else "")
        status_s = str(status).strip() if status is not None else ""

        if req_s and status_s and status_s.lower() != "satisfied":
            if req_s not in seen_unsat:
                seen_unsat.add(req_s)
                unsat.append(req_s)

        # Extract course codes from "Registrations Used" if present.
        if isinstance(regs, str) and regs.strip():
            m = _COURSE_CODE_RE.search(regs.upper())
            if m:
                completed.add(_norm_code(f"{m.group(1)} {m.group(2)}"))

    return AcademicProgress(completed_codes=completed, unsatisfied_requirements=unsat)

