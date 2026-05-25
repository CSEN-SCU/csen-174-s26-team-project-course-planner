from __future__ import annotations

import re
from collections import defaultdict
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


def _merge_requirement_statuses(status_set: set[str]) -> str:
    if "Not Satisfied" in status_set:
        return "Not Satisfied"
    if "In Progress" in status_set:
        return "In Progress"
    return next(iter(sorted(status_set))) if status_set else ""

_CELL_NOISE_RE = (
    # Parenthetical fragments in titles to strip before parsing the course code (code is in the first clause)
    re.compile(r"\s*\([^)]*In Progress[^)]*\)\s*", re.IGNORECASE),
    re.compile(r"\s*\([^)]*Transfer Credit[^)]*\)\s*", re.IGNORECASE),
)


def sanitize_parsed_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Return a copy of parsed progress rows with any ``grade`` field removed."""
    out: list[dict[str, Any]] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        cleaned = {k: v for k, v in row.items() if k != "grade"}
        out.append(cleaned)
    return out


def registration_to_course_code(cell: str | None) -> str | None:
    """Extract a course code like COEN 10 / CSEN 140L from cells shaped like ``COEN 10 - Introduction …``."""
    if cell is None or not isinstance(cell, str):
        return None
    head = cell.split(" - ", 1)[0].strip()
    for rx in _CELL_NOISE_RE:
        head = rx.sub("", head).strip()
    parts = head.split()
    if len(parts) < 2:
        return None
    subj = parts[0].strip().upper()
    num = parts[1].strip().upper()
    if len(subj) < 2 or len(subj) > 8:
        return None
    if not re.fullmatch(r"[A-Z]{2,8}", subj):
        return None
    if not re.fullmatch(r"\d+[A-Z]*", num):
        return None
    return f"{subj} {num}"


def _column_map(header_row: tuple[Any, ...]) -> dict[str, int] | None:
    """Map Workday header labels to column indices (layout varies by export)."""
    headers = [str(c).strip() if c is not None else "" for c in header_row]

    def _idx(*names: str) -> int | None:
        for name in names:
            if name in headers:
                return headers.index(name)
        return None

    req = _idx("Requirement")
    if req is None:
        return None
    status = _idx("Status")
    return {
        "requirement": req,
        "status": status if status is not None else req + 1,
        "remaining": _idx("Remaining"),
        "registration": _idx("Registration", "Registrations Used", "Registrations"),
        "period": _idx("Academic Period", "Period"),
        "units": _idx("Units"),
    }


def _cell(row: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index >= len(row):
        return None
    return row[index]


def _find_progress_sheet(wb: Any) -> tuple[Any, dict[str, int] | None]:
    """Return the worksheet and column map for the first sheet with a Requirement header."""
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(max_row=40, values_only=True):
            if not row:
                continue
            colmap = _column_map(row)
            if colmap is not None:
                return ws, colmap
    return wb.active, None


def parse_academic_progress_xlsx(xlsx_bytes: bytes) -> dict[str, Any]:
    """Parse SCU View My Academic Progress export (single-column sheet layout).

    ``detail_rows`` is suitable for ``st.dataframe``;
    ``not_satisfied`` summarizes blocks still Not Satisfied;
    ``course_codes`` lists all parsed codes from registration rows (unique, sorted).
    """
    # Workday exports break openpyxl read_only (rows collapse to a single cell); use normal load.
    wb = load_workbook(BytesIO(xlsx_bytes), read_only=False, data_only=True)
    detail_rows: list[dict[str, Any]] = []
    not_satisfied: list[dict[str, Any]] = []
    all_codes: list[str] = []
    requirement_status_sets: defaultdict[str, set[str]] = defaultdict(set)
    requirement_status: dict[str, str] = {}

    try:
        ws, colmap = _find_progress_sheet(wb)
        if colmap is None:
            return {
                "detail_rows": [],
                "not_satisfied": [],
                "course_codes": [],
                "requirement_status": {},
                "requirement_status_counts": {},
            }

        header_found = False
        for row in ws.iter_rows(values_only=True):
            if not header_found:
                if _column_map(row) is not None:
                    header_found = True
                continue
            if row is None or all(c is None or str(c).strip() == "" for c in row[:4]):
                continue

            requirement = _cell(row, colmap["requirement"])
            status = _cell(row, colmap["status"])
            remaining = _cell(row, colmap["remaining"])
            registration = _cell(row, colmap["registration"])
            period = _cell(row, colmap["period"])
            units = _cell(row, colmap["units"])
            # Grade column is intentionally ignored — never stored or returned.

            rq = str(requirement).strip() if requirement is not None else ""
            if not rq:
                continue
            st = str(status).strip()
            rm = remaining if remaining is None else (
                remaining if isinstance(remaining, (int, float)) else str(remaining).strip() or None
            )
            reg = registration if registration is None else (
                str(registration).strip() or None
            )

            if st:
                requirement_status_sets[rq].add(st)

            code = registration_to_course_code(reg)
            if code:
                all_codes.append(code)

            detail_rows.append(
                {
                    "requirement": rq,
                    "status": st,
                    "remaining": rm,
                    "registration": reg,
                    "course_code": code,
                    "academic_period": period,
                    "units": units,
                }
            )

        requirement_status = {
            rq: _merge_requirement_statuses(seen)
            for rq, seen in sorted(requirement_status_sets.items())
        }

        for rq, merged in requirement_status.items():
            if merged == "Not Satisfied":
                exemplar = next(
                    (
                        r
                        for r in detail_rows
                        if r["requirement"] == rq and r["status"] == "Not Satisfied"
                    ),
                    next((r for r in detail_rows if r["requirement"] == rq), {}),
                )
                not_satisfied.append(
                    {
                        "requirement": rq,
                        "remaining": exemplar.get("remaining"),
                        "status": "Not Satisfied",
                    }
                )

    finally:
        wb.close()

    course_codes = sorted(set(all_codes), key=lambda c: (c.split()[0], c.split()[1] if len(c.split()) > 1 else ""))

    stats: dict[str, int] = {}
    for rq, merged in requirement_status.items():
        stats[merged] = stats.get(merged, 0) + 1

    return {
        "detail_rows": detail_rows,
        "not_satisfied": not_satisfied,
        "course_codes": course_codes,
        "requirement_status": requirement_status,
        "requirement_status_counts": stats,
    }
