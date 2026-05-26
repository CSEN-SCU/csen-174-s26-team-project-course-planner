"""
Parse SCU Find Course / Sections-style exports (.xlsx) to align recommended courses with
scheduled instructors AND real meeting days/times.

Default files (relative to the ``course_planner/`` package directory):
- ``SCU_Find_Course_Sections.xlsx``
- ``scu_find_course.xlsx``

Index entry shape:
  {
    "instructors": list[str],          # unique instructor names
    "meeting_days": list[int],         # 0=Mon … 4=Fri
    "meeting_start_min": int | None,   # minutes from 8:00 AM
    "meeting_end_min":   int | None,   # minutes from 8:00 AM
  }
"""

from __future__ import annotations

import csv
import re
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

_COURSE_PLANNER_DIR = Path(__file__).resolve().parents[1]
_DEFAULT_SCHEDULE_FILES = (
    _COURSE_PLANNER_DIR / "SCU_Find_Course_Sections.xlsx",
    _COURSE_PLANNER_DIR / "scu_find_course.xlsx",
)
_INSTRUCTOR_RATINGS_FILE = _COURSE_PLANNER_DIR / "data" / "instructor_ratings.csv"

_SCHEDULE_SUBJECT_TYPOS: dict[str, str] = {"CSEE": "CSEN"}

_CALENDAR_START_MIN = 8 * 60   # 8:00 AM
_CALENDAR_END_MIN   = 18 * 60  # 6:00 PM

# Day token → weekday index (0=Mon, 4=Fri)
_DAY_TOKEN_MAP: dict[str, int] = {
    "M": 0, "MON": 0, "MONDAY": 0,
    "T": 1, "TU": 1, "TUE": 1, "TUES": 1, "TUESDAY": 1,
    "W": 2, "WED": 2, "WEDNESDAY": 2,
    "TH": 3, "THU": 3, "THUR": 3, "THURS": 3, "THURSDAY": 3, "R": 3,
    "F": 4, "FRI": 4, "FRIDAY": 4,
}

# Candidate column header names (lower-cased for matching)
_DAYS_HEADERS   = {"days", "day", "meeting days", "meeting day", "mtg days", "mtg day"}
_START_HEADERS  = {"mtg start", "meeting start", "start time", "start", "begin time", "begin"}
_END_HEADERS    = {"mtg end", "meeting end", "end time", "end"}
_TIMES_HEADERS  = {"times", "meeting times", "time", "meeting time", "mtg time", "meeting patterns", "meeting pattern", "mtg patterns", "patterns"}
_TAGS_HEADERS   = {"course tags", "tags", "categories", "category", "course categories"}


# ── helpers ─────────────────────────────────────────────────────────────────

def expand_subjects_for_schedule_lookup(subject_tokens: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for raw in subject_tokens:
        u = raw.strip().upper()
        if not u:
            continue
        for cand in (u, _SCHEDULE_SUBJECT_TYPOS.get(u, "")):
            if cand and cand not in seen:
                out.append(cand)
                seen.add(cand)
    return out


def _find_schedule_path(explicit: Path | None) -> Path | None:
    if explicit is not None and explicit.is_file():
        return explicit
    for p in _DEFAULT_SCHEDULE_FILES:
        if p.is_file():
            return p
    return None


def _parse_section_subject_number(
    course_section: str | None,
) -> tuple[str, str] | None:
    if not course_section or not isinstance(course_section, str):
        return None
    head = course_section.split(" - ")[0].strip().upper()
    m = re.match(r"^([A-Z]{2,8})\s+(\d+[A-Z]?)\s*-\s*\d+\s*$", head)
    if not m:
        return None
    return m.group(1), m.group(2)


def _parse_section_subject_number_with_sec(
    course_section: str | None,
) -> tuple[str, str, int] | None:
    """Same as above but also returns the section number (e.g. '1' from 'CSEN 194L-1')."""
    if not course_section or not isinstance(course_section, str):
        return None
    head = course_section.split(" - ")[0].strip().upper()
    m = re.match(r"^([A-Z]{2,8})\s+(\d+[A-Z]?)\s*-\s*(\d+)\s*$", head)
    if not m:
        return None
    return m.group(1), m.group(2), int(m.group(3))


def _parse_days(cell: Any) -> list[int]:
    """'M W F' or 'MWF' or 'M,W,F' → [0, 2, 4]."""
    if cell is None:
        return []
    text = str(cell).upper().strip()
    # Remove anything after "|" (in case days+times are in one cell, e.g. "M W F | 9:15 AM")
    text = text.split("|")[0].strip()
    # Normalise separators
    text = text.replace(",", " ").replace("/", " ")
    # Handle compact "MWF" / "MW" without spaces by inserting spaces between known tokens
    # (longest first to avoid "TH" being consumed as "T" + "H")
    expanded = re.sub(r"\b(TH|MON|TUE|TUES|WED|THU|THUR|THURS|FRI|TU)\b", r" \1 ", text, flags=re.I)
    expanded = re.sub(r"\b([MTWRF])\b", r" \1 ", expanded)
    days: list[int] = []
    for tok in expanded.split():
        idx = _DAY_TOKEN_MAP.get(tok.upper())
        if idx is not None and idx not in days:
            days.append(idx)
    return sorted(days)


def _parse_single_time(s: str) -> int | None:
    """'9:15 AM' → minutes from midnight."""
    m = re.fullmatch(r"(\d{1,2}):(\d{2})\s*(AM|PM)?", s.strip(), re.IGNORECASE)
    if not m:
        return None
    h, mn = int(m.group(1)), int(m.group(2))
    ampm = (m.group(3) or "").upper()
    if ampm == "PM" and h != 12:
        h += 12
    elif ampm == "AM" and h == 12:
        h = 0
    return h * 60 + mn


def _offset(total_min: int) -> int:
    """Minutes-from-midnight → minutes-from-8AM, clamped to calendar range."""
    return max(0, min(_CALENDAR_END_MIN - _CALENDAR_START_MIN, total_min - _CALENDAR_START_MIN))


def _parse_time_range(cell: Any) -> tuple[int, int] | None:
    """'9:15 AM - 10:20 AM' → (start_offset, end_offset) in minutes from 8 AM."""
    if cell is None:
        return None
    text = str(cell).strip()
    # Strip leading days portion if combined: "M W F | 9:15 AM - 10:20 AM"
    if "|" in text:
        text = text.split("|", 1)[1].strip()
    m = re.search(
        r"(\d{1,2}:\d{2}\s*(?:AM|PM)?)\s*[-–]\s*(\d{1,2}:\d{2}\s*(?:AM|PM)?)",
        text,
        re.IGNORECASE,
    )
    if not m:
        return None
    s = _parse_single_time(m.group(1))
    e = _parse_single_time(m.group(2))
    if s is None or e is None:
        return None
    # If PM marker absent on start, inherit from end when end > 12:00 noon
    if e >= 12 * 60 and s < 12 * 60 and "pm" not in m.group(1).lower() and "am" not in m.group(1).lower():
        # Both times likely PM (e.g. "1:00 - 2:00 PM")
        pass  # leave as-is; single-time parser already handles explicit AM/PM
    start_off = _offset(s)
    end_off   = _offset(e)
    if start_off >= end_off:
        return None
    return start_off, end_off


def _normalize_planner_course_text(course_code: str) -> str:
    u = course_code.upper().replace("&", " ").replace(",", " ")
    u = re.sub(r"(\d+)\s*/\s*L\b", r"\1L", u)
    u = u.replace("/", " ")
    return " ".join(u.split())


def planned_section_keys(course_code: str) -> set[tuple[str, str]]:
    text = _normalize_planner_course_text(course_code)
    tokens = [t for t in text.split() if t]
    keys: set[tuple[str, str]] = set()
    i = 0
    while i < len(tokens):
        t = tokens[i]
        if (
            i + 1 < len(tokens)
            and re.fullmatch(r"[A-Z]{2,8}", t)
            and re.fullmatch(r"\d+[A-Z]?", tokens[i + 1])
        ):
            subj, num = t, tokens[i + 1]
            keys.add((subj, num))
            typo = _SCHEDULE_SUBJECT_TYPOS.get(subj)
            if typo:
                keys.add((typo, num))
                subj = typo
            if subj == "COEN":
                keys.add(("CSEN", num))
            elif subj == "CSEN":
                keys.add(("COEN", num))
            elif subj == "ECEN":
                keys.add(("ELEN", num))
            elif subj == "ELEN":
                keys.add(("ECEN", num))
            i += 2
            continue
        i += 1
    return keys


def _split_instructor_aliases(cell: Any) -> list[str]:
    if not cell or not isinstance(cell, str):
        return []
    return [p.strip() for p in re.split(r"\|", cell) if p.strip() and p.strip().lower() != "none"]


def _find_col(header: list[str], candidates: set[str]) -> int | None:
    for i, h in enumerate(header):
        if h.strip().lower() in candidates:
            return i
    return None


def load_schedule_section_index(path: Path | None = None) -> dict[tuple[str, str], dict[str, Any]]:
    """
    Read xlsx and build (subject, catalog_number) -> entry dict:
      {instructors, meeting_days, meeting_start_min, meeting_end_min}
    """
    p = _find_schedule_path(path)
    if p is None:
        return {}

    index: dict[tuple[str, str], dict[str, Any]] = {}
    wb = load_workbook(p, read_only=True, data_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header_row = next(it, None)
        if not header_row:
            return {}
        h = [str(c).strip() if c is not None else "" for c in header_row]

        # Required columns
        try:
            idx_sec = h.index("Course Section")
        except ValueError:
            return {}
        idx_inst = next((i for i, x in enumerate(h) if x == "All Instructors"), None)

        # Optional time columns
        idx_days  = _find_col(h, _DAYS_HEADERS)
        idx_start = _find_col(h, _START_HEADERS)
        idx_end   = _find_col(h, _END_HEADERS)
        idx_times = _find_col(h, _TIMES_HEADERS)  # combined "9:15 AM - 10:20 AM"

        def _get(row: tuple, idx: int | None) -> Any:
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        for row in it:
            if not row or idx_sec >= len(row):
                continue
            key = _parse_section_subject_number(row[idx_sec])
            if not key:
                continue

            names = _split_instructor_aliases(_get(row, idx_inst)) if idx_inst is not None else []

            # Parse meeting days
            days: list[int] = []
            raw_days = _get(row, idx_days)
            if raw_days:
                days = _parse_days(raw_days)
            elif idx_times is not None:
                # Combined cell may contain "M W F | 9:15 AM - 10:20 AM"
                days = _parse_days(_get(row, idx_times))

            # Parse meeting times
            time_range: tuple[int, int] | None = None
            if idx_start is not None and idx_end is not None:
                raw_s = _get(row, idx_start)
                raw_e = _get(row, idx_end)
                if raw_s and raw_e:
                    s = _parse_single_time(str(raw_s))
                    e = _parse_single_time(str(raw_e))
                    if s is not None and e is not None:
                        off_s = _offset(s)
                        off_e = _offset(e)
                        if off_s < off_e:
                            time_range = (off_s, off_e)
            if time_range is None and idx_times is not None:
                time_range = _parse_time_range(_get(row, idx_times))
            if time_range is None and idx_days is None and idx_times is not None:
                # fallback: try combined days+times cell
                time_range = _parse_time_range(_get(row, idx_times))

            entry = index.setdefault(
                key,
                {"instructors": [], "meeting_days": [], "meeting_start_min": None, "meeting_end_min": None},
            )
            for n in names:
                if n not in entry["instructors"]:
                    entry["instructors"].append(n)
            if days and not entry["meeting_days"]:
                entry["meeting_days"] = days
            if time_range and entry["meeting_start_min"] is None:
                entry["meeting_start_min"] = time_range[0]
                entry["meeting_end_min"] = time_range[1]

    finally:
        wb.close()

    _merge_lab_instructors_into_base(index)
    _mirror_ecen_elen_keys(index)
    return index


def _merge_lab_instructors_into_base(index: dict[tuple[str, str], dict[str, Any]]) -> None:
    for (subj, num) in list(index.keys()):
        s = str(num)
        if not s.endswith("L") or len(s) < 2 or not s[:-1].isdigit():
            continue
        base_key = (subj, s[:-1])
        base = index.setdefault(base_key, {"instructors": [], "meeting_days": [], "meeting_start_min": None, "meeting_end_min": None})
        for n in index.get((subj, s), {}).get("instructors", []):
            if n not in base["instructors"]:
                base["instructors"].append(n)


def _mirror_ecen_elen_keys(index: dict[tuple[str, str], dict[str, Any]]) -> None:
    for (subj, num) in list(index.keys()):
        if subj == "ECEN":
            alt = ("ELEN", num)
            if alt not in index:
                index[alt] = index[(subj, num)]
        elif subj == "ELEN":
            alt = ("ECEN", num)
            if alt not in index:
                index[alt] = index[(subj, num)]


def scheduled_instructors_for_course(
    course_code: str, index: dict[tuple[str, str], dict[str, Any]]
) -> list[str]:
    if not index:
        return []
    want = planned_section_keys(course_code)
    out: list[str] = []
    for k in want:
        entry = index.get(k)
        if entry:
            for name in entry.get("instructors", []):
                if name not in out:
                    out.append(name)
    return out


def meeting_times_for_course(
    course_code: str, index: dict[tuple[str, str], dict[str, Any]]
) -> dict[str, Any] | None:
    """Return {meeting_days, meeting_start_min, meeting_end_min} or None if not found."""
    if not index:
        return None
    for k in planned_section_keys(course_code):
        entry = index.get(k)
        if entry and entry.get("meeting_days") and entry.get("meeting_start_min") is not None:
            return {
                "meeting_days": entry["meeting_days"],
                "meeting_start_min": entry["meeting_start_min"],
                "meeting_end_min": entry["meeting_end_min"],
            }
    return None


def load_all_course_sections(
    path: "Path | None" = None,
) -> "dict[tuple[str, str], list[dict[str, Any]]]":
    """
    Return every lab/lecture section row as a list per (subject, number).
    Each entry: {section, meeting_days, meeting_start_min, meeting_end_min, instructors}
    Useful for displaying all available lab sections to the student.
    """
    p = _find_schedule_path(path)
    if p is None:
        return {}

    result: dict[tuple[str, str], list[dict[str, Any]]] = {}
    wb = load_workbook(p, read_only=True, data_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header_row = next(it, None)
        if not header_row:
            return {}
        h = [str(c).strip() if c is not None else "" for c in header_row]

        try:
            idx_sec = h.index("Course Section")
        except ValueError:
            return {}
        idx_inst = next((i for i, x in enumerate(h) if x == "All Instructors"), None)
        idx_days  = _find_col(h, _DAYS_HEADERS)
        idx_start = _find_col(h, _START_HEADERS)
        idx_end   = _find_col(h, _END_HEADERS)
        idx_times = _find_col(h, _TIMES_HEADERS)

        def _get(row: tuple, idx: "int | None") -> Any:
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        for row in it:
            if not row or idx_sec >= len(row):
                continue
            parsed = _parse_section_subject_number_with_sec(row[idx_sec])
            if not parsed:
                continue
            subj, num, sec_num = parsed

            names = _split_instructor_aliases(_get(row, idx_inst)) if idx_inst is not None else []

            days: list[int] = []
            raw_days = _get(row, idx_days)
            if raw_days:
                days = _parse_days(raw_days)
            elif idx_times is not None:
                days = _parse_days(_get(row, idx_times))

            time_range: tuple[int, int] | None = None
            if idx_start is not None and idx_end is not None:
                raw_s = _get(row, idx_start)
                raw_e = _get(row, idx_end)
                if raw_s and raw_e:
                    s = _parse_single_time(str(raw_s))
                    e = _parse_single_time(str(raw_e))
                    if s is not None and e is not None:
                        off_s = _offset(s)
                        off_e = _offset(e)
                        if off_s < off_e:
                            time_range = (off_s, off_e)
            if time_range is None and idx_times is not None:
                time_range = _parse_time_range(_get(row, idx_times))

            key = (subj, num)
            result.setdefault(key, []).append({
                "section": sec_num,
                "meeting_days": days,
                "meeting_start_min": time_range[0] if time_range else None,
                "meeting_end_min": time_range[1] if time_range else None,
                "instructors": names,
            })

        # Mirror ECEN ↔ ELEN
        for (subj, num) in list(result.keys()):
            if subj == "ECEN":
                result.setdefault(("ELEN", num), result[(subj, num)])
            elif subj == "ELEN":
                result.setdefault(("ECEN", num), result[(subj, num)])

    finally:
        wb.close()

    return result


def _parse_course_tag_codes(tags_cell: Any) -> list[str]:
    """Extract normalised tag strings from a 'Course Tags' cell.

    Each line is one tag in the format:
        "Tag Group :: Short Code | Long Description"

    We return both the short code and the long description so callers
    can match against either.  Example input:
        "Core Explorations :: RTC 3 | Religion, Theology and Culture 3\\n\\n
         Core Integrations :: ELSJ | Experiential Learning for Social Justice"
    → ["RTC 3", "Religion, Theology and Culture 3",
       "ELSJ", "Experiential Learning for Social Justice"]
    """
    if not tags_cell:
        return []
    results: list[str] = []
    for line in str(tags_cell).split("\n"):
        line = line.strip()
        if not line:
            continue
        # Part after "::" (or the whole line if no "::")
        part = line.split("::", 1)[1].strip() if "::" in line else line
        # Split short code from long description
        if "|" in part:
            short, long = part.split("|", 1)
            short = short.strip()
            long = long.strip()
            if short:
                results.append(short)
            if long:
                results.append(long)
        elif part:
            results.append(part)
    return results


def load_category_course_index(path: Path | None = None) -> dict[str, list[str]]:
    """Build a reverse mapping: normalised_tag_text → [course_code, ...].

    Keys are lower-cased tag texts (both short codes like "RTC 3" and long
    descriptions like "Religion, Theology and Culture 3").

    Used in the planning agent to find courses that satisfy open Core / GE
    requirements that have no explicit course code in the Workday transcript
    (e.g. "Core: ENGR: RTC 3", "Core: ENGR: Experiential Learning for Social Justice").
    """
    p = _find_schedule_path(path)
    if p is None:
        return {}

    index: dict[str, list[str]] = {}
    wb = load_workbook(p, read_only=True, data_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header_row = next(it, None)
        if not header_row:
            return {}
        h = [str(c).strip() if c is not None else "" for c in header_row]

        try:
            idx_sec = h.index("Course Section")
        except ValueError:
            return {}
        idx_tags = _find_col(h, _TAGS_HEADERS)
        if idx_tags is None:
            return {}

        def _get(row: tuple, idx: int | None) -> Any:
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        for row in it:
            if not row or idx_sec >= len(row):
                continue
            key = _parse_section_subject_number(row[idx_sec])
            if not key:
                continue
            subj, num = key
            course_code = f"{subj} {num}"

            for tag in _parse_course_tag_codes(_get(row, idx_tags)):
                norm = tag.strip().lower()
                if not norm:
                    continue
                bucket = index.setdefault(norm, [])
                if course_code not in bucket:
                    bucket.append(course_code)
                # Add CSEN ↔ COEN alias
                if subj == "CSEN":
                    alt = f"COEN {num}"
                    if alt not in bucket:
                        bucket.append(alt)
                elif subj == "COEN":
                    alt = f"CSEN {num}"
                    if alt not in bucket:
                        bucket.append(alt)
    finally:
        wb.close()

    return index


@lru_cache(maxsize=1)
def load_core_integrations_course_set(path: str | None = None) -> frozenset[str]:
    """Return the set of course codes that carry at least one 'Core Integrations ::' tag.

    Used by R4 to restrict Educational Enrichment candidates to courses that
    actually carry a Core Integrations tag, not the broader Pathways pool.
    Returns ``frozenset()`` when the schedule xlsx is absent.
    """
    p = _find_schedule_path(Path(path) if path else None)
    if p is None:
        return frozenset()

    codes: set[str] = set()
    wb = load_workbook(p, read_only=True, data_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header_row = next(it, None)
        if not header_row:
            return frozenset()
        h = [str(c).strip() if c is not None else "" for c in header_row]
        try:
            idx_sec = h.index("Course Section")
        except ValueError:
            return frozenset()
        idx_tags = _find_col(h, _TAGS_HEADERS)
        if idx_tags is None:
            return frozenset()

        for row in it:
            if not row or idx_sec >= len(row):
                continue
            key = _parse_section_subject_number(row[idx_sec])
            if not key:
                continue
            subj, num = key
            tags_cell = row[idx_tags] if idx_tags < len(row) else None
            if not tags_cell:
                continue
            # A course qualifies if any tag line starts with "Core Integrations ::"
            for line in str(tags_cell).split("\n"):
                if line.strip().lower().startswith("core integrations ::"):
                    course_code = f"{subj} {num}"
                    codes.add(course_code)
                    if subj == "CSEN":
                        codes.add(f"COEN {num}")
                    elif subj == "COEN":
                        codes.add(f"CSEN {num}")
                    break
    finally:
        wb.close()

    return frozenset(codes)


def load_course_titles_index(path: Path | None = None) -> dict[tuple[str, str], str]:
    """Build a (subject, number) → canonical course title index.

    Reads the 'Course Section' column of the schedule xlsx, which has the
    form ``"CSEN 122-1 - Computer Architecture (-)"``, and returns the
    portion after ``" - "`` (stripped of trailing parenthetical tags like
    ``"(-)"`` or ``"(In Progress)"``).

    Used to OVERRIDE the title field on courses recommended by the LLM,
    so we don't ship hallucinated names to the UI (e.g. an LLM labelling
    CSEN 122L as "Data Structures and Algorithms Lab" — the schedule
    truth is "Computer Architecture Laboratory").
    """
    p = _find_schedule_path(path)
    if p is None:
        return {}
    index: dict[tuple[str, str], str] = {}
    wb = load_workbook(p, read_only=True, data_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if not header:
            return {}
        for row in it:
            if not row or not row[0]:
                continue
            key = _parse_section_subject_number(row[0])
            if not key:
                continue
            cell = str(row[0])
            _, _, tail = cell.partition(" - ")
            if not tail:
                continue
            # Strip trailing "(...)" tags such as "(-)", "(In Progress)"
            title = re.sub(r"\s*\([^)]*\)\s*$", "", tail).strip()
            if not title:
                continue
            if key not in index:
                index[key] = title
        # Mirror CSEN ↔ COEN and ECEN ↔ ELEN so either alias resolves
        for (subj, num) in list(index.keys()):
            if subj == "CSEN":
                index.setdefault(("COEN", num), index[(subj, num)])
            elif subj == "COEN":
                index.setdefault(("CSEN", num), index[(subj, num)])
            elif subj == "ECEN":
                index.setdefault(("ELEN", num), index[(subj, num)])
            elif subj == "ELEN":
                index.setdefault(("ECEN", num), index[(subj, num)])
    finally:
        wb.close()
    return index


def course_title_for(course_code: str, titles_index: dict[tuple[str, str], str]) -> str | None:
    """Return canonical title for ``course_code``, trying CSEN/COEN aliases."""
    if not titles_index:
        return None
    for key in planned_section_keys(course_code):
        title = titles_index.get(key)
        if title:
            return title
    return None


def load_course_units_index(path: Path | None = None) -> dict[tuple[str, str], int]:
    """Build a (subject, number) → units index from the schedule xlsx 'Units'
    column.

    Used to OVERRIDE the units field on courses recommended by the LLM. The
    model frequently invents unit counts (e.g. CSEN 122 as 3u, its lab as 2u);
    the catalog truth is CSEN 122 = 4u, CSEN 122L = 1u. CSEN↔COEN and
    ECEN↔ELEN aliases are mirrored.
    """
    p = _find_schedule_path(path)
    if p is None:
        return {}
    index: dict[tuple[str, str], int] = {}
    wb = load_workbook(p, read_only=True, data_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header_row = next(it, None)
        if not header_row:
            return {}
        h = [str(c).strip() if c is not None else "" for c in header_row]
        try:
            idx_sec = h.index("Course Section")
        except ValueError:
            return {}
        idx_units = next((i for i, x in enumerate(h) if x == "Units"), None)
        if idx_units is None:
            return {}
        for row in it:
            if not row or idx_sec >= len(row):
                continue
            key = _parse_section_subject_number(row[idx_sec])
            if not key:
                continue
            raw = row[idx_units] if idx_units < len(row) else None
            try:
                units = int(float(str(raw).strip()))
            except (TypeError, ValueError):
                continue
            if key not in index:
                index[key] = units
        for (subj, num) in list(index.keys()):
            if subj == "CSEN":
                index.setdefault(("COEN", num), index[(subj, num)])
            elif subj == "COEN":
                index.setdefault(("CSEN", num), index[(subj, num)])
            elif subj == "ECEN":
                index.setdefault(("ELEN", num), index[(subj, num)])
            elif subj == "ELEN":
                index.setdefault(("ECEN", num), index[(subj, num)])
    finally:
        wb.close()
    return index


def course_units_for(course_code: str, units_index: dict[tuple[str, str], int]) -> int | None:
    """Return catalog units for ``course_code``, trying CSEN/COEN aliases."""
    if not units_index:
        return None
    for key in planned_section_keys(course_code):
        u = units_index.get(key)
        if u is not None:
            return u
    return None


# ── Instructor ratings (R5) ──────────────────────────────────────────────────


def _coerce_float(s: str | None) -> float | None:
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


@lru_cache(maxsize=1)
def load_instructor_ratings(
    path: str | None = None,
) -> dict[str, dict[str, Any]]:
    """Load ``data/instructor_ratings.csv`` into ``{name: rating_dict}``.

    Lines beginning with ``#`` are treated as comments and skipped (the file
    carries a provenance header). Names are stored verbatim AND under a
    lower-cased key so lookups are case-insensitive.

    Returns ``{}`` if the file is absent. Rows with an unparseable rating get
    ``rating=None`` so the picker can fall back gracefully.
    """
    p = Path(path) if path else _INSTRUCTOR_RATINGS_FILE
    if not p.is_file():
        return {}
    out: dict[str, dict[str, Any]] = {}
    with p.open("r", encoding="utf-8") as fh:
        lines = [ln for ln in fh if not ln.lstrip().startswith("#")]
    reader = csv.DictReader(lines)
    for row in reader:
        name = (row.get("instructor_name") or "").strip()
        if not name:
            continue
        rec = {
            "instructor": name,
            "rating": _coerce_float(row.get("rating")),
            "difficulty": _coerce_float(row.get("difficulty")),
            "would_take_again_pct": _coerce_float(row.get("would_take_again_pct")),
            "source": (row.get("source") or "unknown").strip(),
        }
        out[name] = rec
        out[name.lower()] = rec
    return out


def instructor_rating_for(
    name: str, ratings: dict[str, dict[str, Any]] | None = None
) -> dict[str, Any]:
    """Return the rating record for ``name``; a stub with ``rating=None`` and
    ``source="unavailable"`` when we have no data for that instructor."""
    if ratings is None:
        ratings = load_instructor_ratings()
    n = (name or "").strip()
    rec = ratings.get(n) or ratings.get(n.lower())
    if rec:
        return dict(rec)
    return {
        "instructor": n,
        "rating": None,
        "difficulty": None,
        "would_take_again_pct": None,
        "source": "unavailable",
    }


def detect_time_conflicts(
    courses: list[str],
    schedule_index: dict[tuple[str, str], dict[str, Any]],
) -> list[tuple[int, int]]:
    """Find pairs of course codes whose meeting times overlap on a shared day.

    Returns a list of ``(idx_a, idx_b)`` index pairs into ``courses`` where
    ``idx_a < idx_b``. A pair conflicts if the two courses share at least
    one weekday and their ``[start, end)`` minute windows overlap.

    Courses with unknown meeting times (no entry in ``schedule_index`` or
    no posted time) cannot conflict and are skipped.
    """
    def _slot(code: str) -> tuple[set[int], int, int] | None:
        for k in planned_section_keys(code):
            entry = schedule_index.get(k)
            if not entry:
                continue
            days = entry.get("meeting_days") or []
            s = entry.get("meeting_start_min")
            e = entry.get("meeting_end_min")
            if days and s is not None and e is not None and s < e:
                return (set(days), int(s), int(e))
        return None

    slots = [_slot(c) for c in courses]
    conflicts: list[tuple[int, int]] = []
    for i, a in enumerate(slots):
        if a is None:
            continue
        for j in range(i + 1, len(slots)):
            b = slots[j]
            if b is None:
                continue
            shared_days = a[0] & b[0]
            if not shared_days:
                continue
            if a[1] < b[2] and b[1] < a[2]:
                conflicts.append((i, j))
    return conflicts


def all_sections_for_course(
    course_code: str,
    sections_index: "dict[tuple[str, str], list[dict[str, Any]]]",
) -> "list[dict[str, Any]]":
    """Return all schedule sections for a course code, combining COEN/CSEN aliases."""
    out: list[dict[str, Any]] = []
    seen_secs: set[int] = set()
    for k in planned_section_keys(course_code):
        for sec in sections_index.get(k, []):
            if sec["section"] not in seen_secs:
                seen_secs.add(sec["section"])
                out.append(sec)
    out.sort(key=lambda s: s["section"])
    return out


# ── Course catalog listing (manual add UI) ──────────────────────────────────

_LAB_PAIR_SUBJECTS = frozenset(
    {"CSEN", "COEN", "CSCI", "ELEN", "ECEN", "PHYS", "CHEM", "BIOL", "MECH"}
)

# Calendar-offset time buckets (minutes from 8:00 AM)
_TIME_BUCKET_RANGES: dict[str, tuple[int, int]] = {
    "morning": (0, 240),
    "afternoon": (240, 540),
    "evening": (540, _CALENDAR_END_MIN - _CALENDAR_START_MIN),
}

_CORE_TAG_PREFIXES = (
    "rtc ",
    "c&i ",
    "elsj",
    "arts",
    "social science",
    "natural science",
    "applied ethics",
    "diversity",
    "advanced writing",
    "critical thinking",
    "values in science",
    "cultures and ideas",
    "religion, theology",
    "experiential learning",
)


def section_overlaps_slot(
    section: dict[str, Any],
    *,
    day_index: int,
    start_min: int,
    end_min: int,
) -> bool:
    """True if section meets on ``day_index`` and meeting time overlaps [start_min, end_min)."""
    meeting_days = section.get("meeting_days") or []
    meeting_start = section.get("meeting_start_min")
    meeting_end = section.get("meeting_end_min")
    if day_index >= 5 or day_index not in meeting_days:
        return False
    if meeting_start is None or meeting_end is None:
        return False
    if meeting_end <= start_min or meeting_start >= end_min:
        return False
    return True


def section_overlaps_time_bucket(section: dict[str, Any], bucket: str) -> bool:
    """True if section meeting time overlaps a named bucket (morning/afternoon/evening)."""
    key = (bucket or "").strip().lower()
    bounds = _TIME_BUCKET_RANGES.get(key)
    if not bounds:
        return True
    meeting_start = section.get("meeting_start_min")
    meeting_end = section.get("meeting_end_min")
    if meeting_start is None or meeting_end is None:
        return False
    b0, b1 = bounds
    return meeting_end > b0 and meeting_start < b1


def _normalize_tag_label(tag: str) -> str:
    return tag.strip().lower()


def _lab_partner_for(subj: str, num: str, sched_keys: set[tuple[str, str]]) -> str | None:
    if subj not in _LAB_PAIR_SUBJECTS:
        return None
    partner_num = num[:-1] if num.endswith("L") else f"{num}L"
    partner_keys = planned_section_keys(f"{subj} {partner_num}")
    if any(k in sched_keys for k in partner_keys):
        return f"{subj} {partner_num}"
    return None


@lru_cache(maxsize=1)
def _cached_offered_sections(path_str: str | None) -> tuple[dict[str, Any], ...]:
    p = Path(path_str) if path_str else None
    return tuple(list_offered_sections(p))


def clear_schedule_caches() -> None:
    """Drop in-process schedule caches after ``SCU_Find_Course_Sections.xlsx`` is replaced.

    ``list_offered_courses`` reads the xlsx on each call, but ``load_core_integrations_course_set``
    and ``load_instructor_ratings`` are ``lru_cache``-d. Pair with
    ``courses._cached_courses.cache_clear()`` in the API router.
    """
    load_core_integrations_course_set.cache_clear()
    load_instructor_ratings.cache_clear()
    _cached_offered_sections.cache_clear()


def list_offered_courses(path: Path | None = None) -> list[dict[str, Any]]:
    """Return every distinct course offered next term, shaped for the manual
    "+ Add course" picker AND for direct placement on the calendar.

    Each entry:
      {course, title, units, professor, meeting_days, meeting_start_min,
       meeting_end_min, lab_partner}

    Meeting fields come from the section index (the representative section);
    ``lab_partner`` is the trailing-L (or de-L) co-requisite code when that
    partner is itself offered, so the frontend can auto-add the pair.
    """
    sched = load_schedule_section_index(path)
    titles = load_course_titles_index(path)
    units = load_course_units_index(path)
    if not sched:
        return []

    # Deduplicate by (subject, number); aliases (COEN mirror of CSEN) are
    # already present as separate keys in the index — collapse them so the
    # picker doesn't show CSEN 122 and COEN 122 as two rows.
    seen: set[tuple[str, str]] = set()
    out: list[dict[str, Any]] = []
    for (subj, num), entry in sched.items():
        # Skip the mirrored alias rows: keep the CSEN/ECEN spelling, drop the
        # COEN/ELEN duplicate when the primary exists.
        if subj in {"COEN", "ELEN"}:
            primary = {"COEN": "CSEN", "ELEN": "ECEN"}[subj]
            if (primary, num) in sched:
                continue
        if (subj, num) in seen:
            continue
        seen.add((subj, num))
        code = f"{subj} {num}"
        instructors = list(entry.get("instructors") or [])
        lab_partner = None
        if subj in _LAB_PAIR_SUBJECTS:
            partner_num = num[:-1] if num.endswith("L") else f"{num}L"
            partner_keys = planned_section_keys(f"{subj} {partner_num}")
            if any(k in sched for k in partner_keys):
                lab_partner = f"{subj} {partner_num}"
        out.append({
            "course": code,
            "title": course_title_for(code, titles),
            "units": course_units_for(code, units),
            "professor": instructors[0] if instructors else None,
            "meeting_days": list(entry.get("meeting_days") or []),
            "meeting_start_min": entry.get("meeting_start_min"),
            "meeting_end_min": entry.get("meeting_end_min"),
            "lab_partner": lab_partner,
        })
    out.sort(key=lambda c: (c["course"].split()[0], c["course"]))
    return out


def list_offered_sections(path: Path | None = None) -> list[dict[str, Any]]:
    """One dict per xlsx row (section) for the manual course browser."""
    p = _find_schedule_path(path)
    if p is None:
        return []

    titles = load_course_titles_index(path)
    units_index = load_course_units_index(path)
    sched_keys: set[tuple[str, str]] = set()
    out: list[dict[str, Any]] = []

    wb = load_workbook(p, read_only=True, data_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header_row = next(it, None)
        if not header_row:
            return []
        h = [str(c).strip() if c is not None else "" for c in header_row]

        try:
            idx_sec = h.index("Course Section")
        except ValueError:
            return []

        idx_subj = h.index("Course Subject") if "Course Subject" in h else None
        idx_num = h.index("Course Number") if "Course Number" in h else None
        idx_secnum = h.index("Section Number") if "Section Number" in h else None
        idx_status = h.index("Section Status") if "Section Status" in h else None
        idx_cap = h.index("Enrolled/Capacity") if "Enrolled/Capacity" in h else None
        idx_inst = next((i for i, x in enumerate(h) if x == "All Instructors"), None)
        idx_units = h.index("Units") if "Units" in h else None
        idx_loc = h.index("Locations") if "Locations" in h else None
        idx_tags = _find_col(h, _TAGS_HEADERS)
        idx_days = _find_col(h, _DAYS_HEADERS)
        idx_start = _find_col(h, _START_HEADERS)
        idx_end = _find_col(h, _END_HEADERS)
        idx_times = _find_col(h, _TIMES_HEADERS)
        idx_patterns = h.index("Meeting Patterns") if "Meeting Patterns" in h else idx_times

        def _get(row: tuple, idx: int | None) -> Any:
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        for row in it:
            if not row or idx_sec >= len(row):
                continue
            parsed = _parse_section_subject_number_with_sec(row[idx_sec])
            if not parsed:
                continue
            subj, num, sec_num = parsed
            sched_keys.add((subj, num))
            code = f"{subj} {num}"

            names = _split_instructor_aliases(_get(row, idx_inst)) if idx_inst is not None else []

            days: list[int] = []
            raw_days = _get(row, idx_days)
            if raw_days:
                days = _parse_days(raw_days)
            elif idx_times is not None:
                days = _parse_days(_get(row, idx_times))

            time_range: tuple[int, int] | None = None
            if idx_start is not None and idx_end is not None:
                raw_s = _get(row, idx_start)
                raw_e = _get(row, idx_end)
                if raw_s and raw_e:
                    s = _parse_single_time(str(raw_s))
                    e = _parse_single_time(str(raw_e))
                    if s is not None and e is not None:
                        off_s = _offset(s)
                        off_e = _offset(e)
                        if off_s < off_e:
                            time_range = (off_s, off_e)
            if time_range is None and idx_times is not None:
                time_range = _parse_time_range(_get(row, idx_times))

            raw_units = _get(row, idx_units)
            try:
                units_val = float(raw_units) if raw_units is not None else None
            except (TypeError, ValueError):
                units_val = course_units_for(code, units_index)

            tags = _parse_course_tag_codes(_get(row, idx_tags))

            meeting_pattern = _get(row, idx_patterns)
            meeting_pattern_str = (
                str(meeting_pattern).strip() if meeting_pattern is not None else ""
            )

            out.append({
                "course_section": str(row[idx_sec]).strip(),
                "course": code,
                "section": int(sec_num) if sec_num is not None else sec_num,
                "subject": str(_get(row, idx_subj) or subj).strip(),
                "number": str(_get(row, idx_num) or num).strip(),
                "title": course_title_for(code, titles),
                "units": units_val if units_val is not None else course_units_for(code, units_index),
                "status": str(_get(row, idx_status) or "").strip() or None,
                "enrolled_capacity": str(_get(row, idx_cap) or "").strip() or None,
                "instructors": names,
                "meeting_days": days,
                "meeting_start_min": time_range[0] if time_range else None,
                "meeting_end_min": time_range[1] if time_range else None,
                "meeting_pattern": meeting_pattern_str or None,
                "location": str(_get(row, idx_loc) or "").strip() or None,
                "course_tags": tags,
                "lab_partner": None,
            })

    finally:
        wb.close()

    for entry in out:
        subj = entry["course"].split()[0]
        num = entry["course"].split()[1] if " " in entry["course"] else ""
        entry["lab_partner"] = _lab_partner_for(subj, num, sched_keys)

    out.sort(key=lambda s: (s["subject"], s["course"], s.get("section") or 0))
    return out


def catalog_facets(sections: list[dict[str, Any]]) -> dict[str, Any]:
    """Build filter facet lists from a section list."""
    subjects: set[str] = set()
    tags_core: set[str] = set()
    tags_other: set[str] = set()

    for s in sections:
        subj = s.get("subject")
        if isinstance(subj, str) and subj.strip():
            subjects.add(subj.strip())
        for tag in s.get("course_tags") or []:
            if not isinstance(tag, str) or not tag.strip():
                continue
            low = tag.lower()
            if any(low.startswith(p) or p in low for p in _CORE_TAG_PREFIXES):
                tags_core.add(tag.strip())
            else:
                tags_other.add(tag.strip())

    return {
        "subjects": sorted(subjects),
        "tags": {
            "Core": sorted(tags_core),
            "Other": sorted(tags_other),
        },
        "time_buckets": list(_TIME_BUCKET_RANGES.keys()),
    }


def filter_catalog_sections(
    sections: list[dict[str, Any]],
    *,
    q: str | None = None,
    subjects: list[str] | None = None,
    days: list[int] | None = None,
    time_buckets: list[str] | None = None,
    tags: list[str] | None = None,
    day_index: int | None = None,
    start_min: int | None = None,
    end_min: int | None = None,
) -> list[dict[str, Any]]:
    """Apply catalog browser filters (AND across filter types, OR within tags/days/buckets)."""
    q_norm = (q or "").strip().lower()
    subject_set = {s.strip().lower() for s in (subjects or []) if s.strip()}
    day_set = set(days or [])
    bucket_list = [b.strip().lower() for b in (time_buckets or []) if b.strip()]
    tag_norms = {_normalize_tag_label(t) for t in (tags or []) if t.strip()}

    out: list[dict[str, Any]] = []
    for s in sections:
        if q_norm:
            hay = " ".join(
                filter(
                    None,
                    [
                        str(s.get("course") or ""),
                        str(s.get("title") or ""),
                        str(s.get("course_section") or ""),
                        " ".join(s.get("instructors") or []),
                    ],
                )
            ).lower()
            if q_norm not in hay:
                continue

        if subject_set:
            subj = str(s.get("subject") or "").strip().lower()
            if subj not in subject_set:
                continue

        if day_set:
            md = set(s.get("meeting_days") or [])
            if not md.intersection(day_set):
                continue

        if bucket_list:
            if not any(section_overlaps_time_bucket(s, b) for b in bucket_list):
                continue

        if tag_norms:
            sec_tags = {_normalize_tag_label(t) for t in (s.get("course_tags") or [])}
            if not sec_tags.intersection(tag_norms):
                continue

        if day_index is not None and start_min is not None and end_min is not None:
            if not section_overlaps_slot(
                s, day_index=day_index, start_min=start_min, end_min=end_min
            ):
                continue

        out.append(s)

    return out
