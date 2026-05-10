import hashlib
import json
import re
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv(Path(__file__).resolve().parent / ".env")

import streamlit as st

from utils.academic_progress_xlsx import parse_academic_progress_xlsx
from utils.meeting_pattern_parse import parse_schedule
from utils.calendar_plan_followup import build_remove_and_replace_preference
from utils.rmp_display import professors_sorted_by_rating
from utils.voice_pref import transcribe_wav_bytes
from utils.scu_course_schedule_xlsx import _parse_section_subject_number
from utils.course_variants import extract_course_variants
from utils.replacement_slot_verify import (
    gap_courses_matching_slot,
    verify_calendar_replacements,
)
from agents.professor_agent import run_professor_agent
from agents.orchestrator import plan_for_user
from agents.requirement_agent import run_requirement_agent
from agents import memory_agent
from auth import streamlit_auth

st.set_page_config(page_title="SCU Course Planner", layout="wide")
st.markdown(
    """
<style>
/* Santa Clara University–inspired UI: off-white main, white chrome, red accents only */

:root {
    --scu-red: #C8102E;
    --scu-red-hover: #A00D26;
    --scu-text: #1A1A1A;
    --scu-heading: #1B263B;
    --scu-bg-main: #F5F5F5;
    --scu-bg-surface: #FFFFFF;
    --scu-border-soft: #E5E5E5;
}

/* Main canvas: light gray */
[data-testid="stAppViewContainer"],
section[data-testid="stMain"],
.main {
    background-color: var(--scu-bg-main) !important;
}

/* Default body copy: charcoal (not red) */
.main,
[data-testid="stSidebar"],
.main p,
.main span,
[data-testid="stMarkdownContainer"] {
    color: var(--scu-text) !important;
}

/* Top toolbar: white + SCU red accent stripe */
[data-testid="stHeader"] {
    background-color: var(--scu-bg-surface) !important;
    border-bottom: 4px solid var(--scu-red) !important;
}

[data-testid="stToolbar"] {
    background-color: var(--scu-bg-surface) !important;
}

/* Headings: dark navy/charcoal — not red */
h1, h2, h3, h4,
.main .block-container h1,
.main .block-container h2,
.main .block-container h3 {
    color: var(--scu-heading) !important;
}

/* Sidebar: warm light gray with red accent bar only */
[data-testid="stSidebar"] {
    background-color: #F8F6F3 !important;
    border-right: 1px solid var(--scu-border-soft) !important;
    box-shadow: inset 4px 0 0 var(--scu-red) !important;
}
[data-testid="stSidebar"] .stMarkdown,
[data-testid="stSidebar"] label {
    color: var(--scu-text) !important;
}

/* Main page title (st.markdown h1) */
h1.scu-main-title {
    font-size: 42px !important;
    color: #1A1A1A !important;
    font-weight: 700 !important;
    margin-bottom: 0.5rem !important;
}

/* Links: SCU red */
a {
    color: var(--scu-red) !important;
}
a:hover {
    color: var(--scu-red-hover) !important;
}

/* Primary buttons: solid red, white label */
.stButton > button {
    background-color: var(--scu-red) !important;
    color: #FFFFFF !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
}
.stButton > button:hover {
    background-color: var(--scu-red-hover) !important;
    color: #FFFFFF !important;
}

/* Metrics: charcoal labels and values */
[data-testid="stMetricLabel"] {
    color: var(--scu-heading) !important;
}
[data-testid="stMetricValue"] {
    color: var(--scu-text) !important;
}

/* Expanders */
.streamlit-expanderHeader {
    color: var(--scu-heading) !important;
    font-weight: 600 !important;
}

/* Card-style containers: white, shadow, red left accent */
[data-testid="stVerticalBlockBorderWrapper"] > div {
    background-color: var(--scu-bg-surface) !important;
    border-left: 3px solid var(--scu-red) !important;
    box-shadow: 0 2px 8px rgba(0, 0, 0, 0.06) !important;
    border-radius: 6px !important;
}

[data-testid="stContainer"] {
    background-color: var(--scu-bg-surface) !important;
    border-left: 3px solid var(--scu-red) !important;
    box-shadow: 0 2px 8px rgba(0, 0, 0, 0.06) !important;
    border-radius: 6px !important;
}

/* Tabs: accent on active */
.stTabs [data-baseweb="tab"][aria-selected="true"] {
    color: var(--scu-red) !important;
    border-bottom-color: var(--scu-red) !important;
}
</style>
""",
    unsafe_allow_html=True,
)

current_user = streamlit_auth.require_login()
if current_user is None:
    st.stop()
USER_ID = int(current_user["id"])


def _try_hydrate_transcript_from_disk(user_id: int) -> None:
    if "missing_details" in st.session_state and "parsed_rows" in st.session_state:
        return
    snap = memory_agent.load_last_transcript_snapshot(user_id)
    if not isinstance(snap, dict):
        return
    md = snap.get("missing_details")
    rows = snap.get("parsed_rows")
    if not isinstance(md, list) or not isinstance(rows, list):
        return
    st.session_state["missing_details"] = md
    st.session_state["parsed_rows"] = rows
    st.session_state["parsed_course_codes"] = list(snap.get("course_codes") or [])
    st.session_state["transcript_progress_snapshot"] = {
        "requirement_status_counts": snap.get("requirement_status_counts") or {},
        "not_satisfied": snap.get("not_satisfied") or [],
        "course_codes": list(snap.get("course_codes") or []),
    }


def _session_transcript_data_for_display() -> dict | None:
    if "parsed_rows" not in st.session_state:
        return None
    snap = st.session_state.get("transcript_progress_snapshot") or {}
    return {
        "detail_rows": list(st.session_state.get("parsed_rows") or []),
        "not_satisfied": snap.get("not_satisfied") or [],
        "course_codes": list(snap.get("course_codes") or []),
        "requirement_status_counts": snap.get("requirement_status_counts") or {},
    }


def _schedule_map_keys_for_row(subject: str, number: str) -> list[str]:
    """Course-code keys aligned with the schedule xlsx, with COEN/CSEN and ECEN/ELEN cross-mapping."""
    subj = subject.strip().upper()
    num = number.strip()
    primary = f"{subj} {num}"
    keys = [primary]
    if subj == "COEN":
        keys.append(f"CSEN {num}")
    elif subj == "CSEN":
        keys.append(f"COEN {num}")
    elif subj == "ECEN":
        keys.append(f"ELEN {num}")
    elif subj == "ELEN":
        keys.append(f"ECEN {num}")
    return keys


def _load_schedule_map_base_from_xlsx() -> dict[str, str]:
    """Build course code -> Meeting Patterns from Find Course Sections xlsx only (no enriched overrides)."""
    base = Path(__file__).resolve().parent
    out: dict[str, str] = {}
    for fname in ("SCU_Find_Course_Sections.xlsx", "scu_find_course.xlsx"):
        p = base / fname
        if not p.is_file():
            continue
        wb = load_workbook(p, read_only=True, data_only=True)
        try:
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            header = next(it, None)
            if not header:
                return out
            h = [str(c).strip() if c is not None else "" for c in header]
            try:
                idx_sec = h.index("Course Section")
                idx_mp = h.index("Meeting Patterns")
            except ValueError:
                return out
            for row in it:
                if not row or max(idx_sec, idx_mp) >= len(row):
                    continue
                sec = row[idx_sec]
                mp = row[idx_mp]
                if mp is None:
                    continue
                mp_s = str(mp).strip()
                if not mp_s:
                    continue
                parsed = _parse_section_subject_number(
                    str(sec).strip() if sec is not None else None
                )
                if not parsed:
                    continue
                subj_s, num_s = parsed[0], parsed[1]
                for k in _schedule_map_keys_for_row(subj_s, num_s):
                    if k not in out:
                        out[k] = mp_s
        finally:
            wb.close()
        break
    return out


def _recommended_item_in_find_course_xlsx(item: dict, base_map: dict[str, str]) -> bool:
    """True if the recommendation matches at least one parseable section in Find Course Sections (variant keys)."""
    if not base_map:
        return True
    course_str = (item.get("course") or "").strip()
    if not course_str:
        return False
    variants = extract_course_variants(course_str)
    for v in variants:
        if " ".join(v.split()).upper() in base_map:
            return True
    return False


def load_course_schedule_map_from_xlsx() -> dict[str, str]:
    """Load Find Course Sections xlsx and build course-code string -> raw Meeting Patterns."""
    out = dict(_load_schedule_map_base_from_xlsx())

    enriched = st.session_state.get("enriched_courses")
    if isinstance(enriched, list):
        for item in enriched:
            course_str = item.get("course") or ""
            if not course_str:
                continue
            variants = extract_course_variants(course_str)
            if not variants:
                continue
            sched = None
            for v in variants:
                key = " ".join(v.split()).upper()
                if key in out:
                    sched = out[key]
                    break
            if sched is None:
                continue
            for v in variants:
                key = " ".join(v.split()).upper()
                out[key] = sched
            out[" ".join(course_str.split())] = sched

    return out


def _norm_course_keys(code: str) -> set[str]:
    out: set[str] = set()
    for v in extract_course_variants(str(code or "")):
        s = " ".join(str(v).split()).upper()
        if s:
            out.add(s)
    return out


def _recommended_fingerprint(recs: list[dict]) -> str:
    return json.dumps(recs, ensure_ascii=False, sort_keys=True, default=str)


def _parse_memory_plan_line(content: str) -> tuple[list[str], int | None]:
    """Extract course codes after PLAN: … and optional total_units from memory ``plan_outcome`` text."""
    text = content or ""
    m = re.search(
        r"PLAN:\s*(.+)\s*\|\s*total_units\s*=\s*(\d+)",
        text,
        re.IGNORECASE | re.DOTALL,
    )
    if m:
        part = m.group(1).strip()
        codes = [c.strip() for c in part.split(",") if c.strip()]
        return codes, int(m.group(2))
    m2 = re.search(r"PLAN:\s*(.+)\s*\|", text, re.IGNORECASE | re.DOTALL)
    if m2:
        part = m2.group(1).strip()
        codes = [c.strip() for c in part.split(",") if c.strip()]
        tu_m = re.search(r"total_units\s*=\s*(\d+)", text, re.IGNORECASE)
        tu = int(tu_m.group(1)) if tu_m else None
        return codes, tu
    m3 = re.search(r"PLAN:\s*(.+)$", text, re.MULTILINE | re.IGNORECASE | re.DOTALL)
    if m3:
        part = m3.group(1).strip()
        part = part.split("\n")[0].strip()
        codes = [c.strip() for c in part.split(",") if c.strip()]
        tu_m = re.search(r"total_units\s*=\s*(\d+)", text, re.IGNORECASE)
        tu = int(tu_m.group(1)) if tu_m else None
        return codes, tu
    return [], None


def _expand_norm_keys_for_gap_match(code: str) -> set[str]:
    """Subject aliases (COEN/CSEN, ECEN/ELEN) so memory codes match DegreeWorks gap rows."""
    keys = set(_norm_course_keys(code))
    parts = " ".join(str(code).split()).split()
    if len(parts) < 2:
        return keys
    subj, tail = parts[0].upper(), " ".join(parts[1:])
    swap = {"COEN": "CSEN", "CSEN": "COEN", "ECEN": "ELEN", "ELEN": "ECEN"}
    if subj in swap:
        keys |= _norm_course_keys(f"{swap[subj]} {tail}")
    return keys


def _units_from_gap_row(row: dict) -> int:
    u = row.get("units")
    if u is None:
        return 0
    try:
        return int(float(u))
    except (TypeError, ValueError):
        return 0


def _match_gap_row_for_restore(parsed_code: str, md: list) -> dict | None:
    want = _expand_norm_keys_for_gap_match(parsed_code)
    if not want:
        return None
    for g in md:
        if not isinstance(g, dict):
            continue
        gc = (g.get("course") or "").strip()
        if not gc:
            continue
        if _expand_norm_keys_for_gap_match(gc) & want:
            return g
    return None


def _recommended_rows_from_parsed_codes(
    codes: list[str],
    missing_details: list | None,
) -> list[dict]:
    """Rebuild ``recommended`` from ``missing_details`` rows matched to parsed PLAN codes."""
    md = missing_details if isinstance(missing_details, list) else []
    out: list[dict] = []
    for code in codes:
        if not (code or "").strip():
            continue
        disp = " ".join(str(code).split())
        matched = _match_gap_row_for_restore(code, md)
        if matched:
            cc = (matched.get("course") or "").strip() or disp
            out.append(
                {
                    "course": cc,
                    "category": (matched.get("category") or "").strip() or "—",
                    "units": _units_from_gap_row(matched),
                    "reason": "Restored from saved plan",
                }
            )
        else:
            out.append(
                {
                    "course": disp,
                    "category": "—",
                    "units": 0,
                    "reason": "Restored from saved plan",
                }
            )
    return out


st.markdown(
    '<h1 class="scu-main-title">SCU Course Planner</h1>',
    unsafe_allow_html=True,
)

COL_LABELS = {
    "requirement": "Requirement / item",
    "status": "Status",
    "remaining": "Remaining (gap)",
    "registration": "Registered course",
    "course_code": "Parsed course code",
    "academic_period": "Term",
    "units": "Units",
    "grade": "Grade",
}

with st.sidebar:
    st.sidebar.markdown(
        """
   <div style='padding: 10px 0 20px 0; border-bottom: 2px solid #C8102E; margin-bottom: 16px;'>
     <span style='color: #C8102E; font-size: 22px; font-weight: 800; letter-spacing: 1px;'>SCU</span>
     <span style='color: #1A1A1A; font-size: 14px; font-weight: 400; margin-left: 8px;'>Course Planner</span>
   </div>
   """,
        unsafe_allow_html=True,
    )
    st.markdown(
        f"Signed in as <span style='font-weight: 700; color: #C8102E;'>{current_user['username']}</span>",
        unsafe_allow_html=True,
    )
    streamlit_auth.logout_button()
    st.divider()
    st.markdown("#### Step 1 · Academic Progress")
    _try_hydrate_transcript_from_disk(USER_ID)
    has_cached_transcript = (
        "missing_details" in st.session_state and "parsed_rows" in st.session_state
    )
    with st.expander(
        "Upload new transcript to replace",
        expanded=not has_cached_transcript,
    ):
        st.caption(
            "`.xlsx` from **SCU → View My Academic Progress** (parsed locally; no API calls)."
        )
        xlsx_file = st.file_uploader(
            "Upload new transcript to replace",
            type=["xlsx"],
            key="academic_progress_xlsx_replace",
        )
        hide_empty = st.checkbox(
            "Hide detail rows that only have status and no registered course",
            value=False,
            key="academic_progress_hide_empty",
        )
        run = st.button("Parse", key="academic_progress_parse_btn")

    st.divider()
    st.markdown("**Curriculum PDF gap analysis** (optional)")
    st.caption(
        "Uses the same **GEMINI_API_KEY** / **GOOGLE_API_KEY** and **GEMINI_MODEL** "
        "as schedule generation. Upload your major-requirements PDF; completed courses "
        "default to codes from your last **Parse** above, plus any you list here."
    )
    pdf_file = st.file_uploader("Major requirements PDF (.pdf)", type=["pdf"])
    pdf_completed_extra = st.text_area(
        "Extra completed course codes (one per line or comma-separated)",
        placeholder="CSEN 174\nCOEN 145",
        key="pdf_completed_extra",
        height=80,
    )
    if st.button("Analyze PDF gaps", type="secondary"):
        if not pdf_file:
            st.warning("Upload a PDF first.")
        else:
            completed = _merge_completed_courses_for_pdf(
                st.session_state.get("parsed_course_codes"), pdf_completed_extra
            )
            if not completed:
                st.warning(
                    "No completed courses to send. Parse Academic Progress above, "
                    "or enter course codes in the extra box."
                )
            else:
                with st.spinner("Analyzing PDF with Gemini…"):
                    try:
                        result = run_requirement_agent(pdf_file.getvalue(), completed)
                    except Exception as e:
                        st.error(f"PDF gap analysis failed: {e}")
                    else:
                        st.session_state["pdf_gap_result"] = result
                        md = _normalize_pdf_missing_details(
                            result.get("missing_details") or []
                        )
                        st.session_state["missing_details"] = md
                        st.success(
                            f"PDF analysis done — {len(md)} gap row(s) ready for Step 2."
                        )
                        st.rerun()

    st.divider()
    with st.expander("My memory", expanded=False):
        st.caption(
            "Past preferences and plans the assistant remembers for *you*. "
            "Delete anything you don't want re-used."
        )
        my_items = memory_agent.list_for_user(USER_ID)
        if not my_items:
            st.info("No memory yet. Generate a recommended schedule to start building one.")
        else:
            for item in my_items:
                with st.container(border=True):
                    st.caption(f"{item['kind']} · {item['created_at']}")
                    st.write(item["content"])
                    if st.button("Restore this plan", key=f"mem_restore_{item['id']}"):
                        codes, tu_mem = _parse_memory_plan_line(item["content"])
                        if not codes:
                            st.warning(
                                "Could not parse PLAN line (expected: PLAN: CSEN 101, … | total_units=…)."
                            )
                        else:
                            md_list = st.session_state.get("missing_details")
                            recs = _recommended_rows_from_parsed_codes(codes, md_list)
                            recomputed = 0
                            for r in recs:
                                try:
                                    recomputed += int(r.get("units") or 0)
                                except (TypeError, ValueError):
                                    pass
                            total_u = tu_mem if tu_mem is not None else recomputed
                            st.session_state["planning_result"] = {
                                "recommended": recs,
                                "total_units": total_u,
                                "advice": "",
                                "assistant_reply": "Restored your saved plan from memory.",
                                "warnings": [],
                            }
                            st.session_state["last_planning_message"] = (
                                "Restored plan from memory"
                            )
                            st.session_state.pop("calendar_replacement_suggestions", None)
                            st.session_state.pop("calendar_replacement_slot_label", None)
                            if recs:
                                with st.spinner("Loading instructor ratings…"):
                                    st.session_state["enriched_courses"] = run_professor_agent(
                                        recs
                                    )
                                st.session_state["_recommended_enrichment_fp"] = (
                                    _recommended_fingerprint(recs)
                                )
                            else:
                                st.session_state.pop("enriched_courses", None)
                                st.session_state.pop("_recommended_enrichment_fp", None)
                            st.success(
                                "Plan restored! Scroll down to see your schedule."
                            )
                            st.rerun()
                    if st.button("Delete", key=f"mem_del_{item['id']}"):
                        memory_agent.delete(USER_ID, item["id"])
                        st.rerun()
            if st.button("Delete all my memory", key="mem_del_all"):
                memory_agent.delete_all_for_user(USER_ID)
                st.rerun()


def _detail_table_rows(rows: list[dict]) -> list[dict]:
    return [{COL_LABELS.get(k, k): row.get(k) for k in COL_LABELS if k in row} for row in rows]


def _not_satisfied_table(items: list[dict]) -> list[dict]:
    return [
        {
            "Requirement / item": i.get("requirement"),
            "Remaining (gap)": i.get("remaining"),
        }
        for i in items
    ]


def _merge_completed_courses_for_pdf(
    parsed_codes: list[str] | None, extra_text: str
) -> list[str]:
    """Union of codes from the last Academic Progress parse plus optional sidebar lines."""
    out: list[str] = []
    seen: set[str] = set()
    for c in parsed_codes or []:
        s = str(c).strip()
        if not s:
            continue
        key = s.upper()
        if key not in seen:
            seen.add(key)
            out.append(s)
    for raw in (extra_text or "").replace(",", "\n").split("\n"):
        s = raw.strip()
        if not s:
            continue
        key = s.upper()
        if key not in seen:
            seen.add(key)
            out.append(s)
    return out


def _normalize_pdf_missing_details(raw: list) -> list[dict]:
    """Keep only rows usable as planner gap items (non-empty course code)."""
    rows: list[dict] = []
    for item in raw or []:
        if not isinstance(item, dict):
            continue
        course = (item.get("course") or "").strip()
        if not course:
            continue
        u = item.get("units", 0)
        try:
            ui = int(u) if u is not None else 0
        except (TypeError, ValueError):
            ui = 0
        cat = (item.get("category") or "").strip() or "(from PDF)"
        rows.append({"course": course, "category": cat, "units": ui})
    return rows


def _rating_display(rating) -> tuple[str, str]:
    """Return (star markdown, numeric text) for st.markdown."""
    if rating is None:
        return "", "—"
    try:
        r = float(rating)
    except (TypeError, ValueError):
        return "", str(rating)
    filled = max(0, min(5, int(round(r))))
    stars = "★" * filled + "☆" * (5 - filled)
    return stars, f"{r:.1f}/5"


def _planning_warning_messages(planning_result: dict) -> list[str]:
    """Human-readable lines from ``planning_agent`` heuristic ``warnings`` (code + message)."""
    raw = planning_result.get("warnings")
    if not isinstance(raw, list):
        return []
    out: list[str] = []
    for w in raw:
        if not isinstance(w, dict):
            continue
        m = (w.get("message") or "").strip()
        if m:
            out.append(m)
    return out


def _augment_gaps_with_removed(
    gaps: list,
    removed_course: str,
    snapshot: dict | None,
) -> list:
    """Copy gaps and append removed course if not already listed (variant match)."""
    removed = (removed_course or "").strip()
    if not removed:
        return list(gaps or [])
    rs = _norm_course_keys(removed)
    for g in gaps or []:
        if not isinstance(g, dict):
            continue
        gc = (g.get("course") or "").strip()
        if gc and rs & _norm_course_keys(gc):
            return list(gaps or [])
    snap = snapshot or {}
    u = snap.get("units", 0)
    try:
        ui = int(u) if u is not None else 0
    except (TypeError, ValueError):
        ui = 0
    cat = (snap.get("category") or "").strip() or "Requirement gap"
    row = {"course": removed, "category": cat, "units": ui}
    return list(gaps or []) + [row]


def _slot_label_from_verify_ctx(verify_ctx: dict | None) -> str:
    if not verify_ctx:
        return ""
    parsed = verify_ctx.get("parsed")
    day = verify_ctx.get("day_label")
    if parsed and day:
        return f"{day} {parsed['start']} – {parsed['end']}"
    if parsed:
        days = ", ".join(str(d) for d in (parsed.get("days") or []) if d)
        return f"{days} {parsed['start']} – {parsed['end']}".strip()
    return "Time TBD"


def _recompute_plan_total_units(plan: dict) -> None:
    total = 0
    for item in plan.get("recommended") or []:
        if not isinstance(item, dict):
            continue
        try:
            total += int(item.get("units") or 0)
        except (TypeError, ValueError):
            continue
    plan["total_units"] = total


def _course_already_in_recommended(plan: dict, course: str) -> bool:
    rs = _norm_course_keys(course)
    for item in plan.get("recommended") or []:
        if not isinstance(item, dict):
            continue
        c = (item.get("course") or "").strip()
        if c and rs & _norm_course_keys(c):
            return True
    return False


def _append_course_to_planning_result(course_row: dict) -> None:
    plan = st.session_state.get("planning_result")
    if not isinstance(plan, dict):
        return
    code = (course_row.get("course") or "").strip()
    if not code or _course_already_in_recommended(plan, code):
        return
    try:
        ui = int(course_row.get("units") or 0)
    except (TypeError, ValueError):
        ui = 0
    plan.setdefault("recommended", []).append(
        {
            "course": code,
            "category": (course_row.get("category") or "").strip() or "—",
            "units": ui,
            "reason": "Added from replacement suggestions for the vacated time slot.",
        }
    )
    _recompute_plan_total_units(plan)
    st.session_state["planning_result"] = plan


def _append_courses_group_to_planning_result(course_rows: list[dict]) -> None:
    for row in course_rows:
        _append_course_to_planning_result(row)


def _lab_partner_course_code(course: str) -> str | None:
    parts = " ".join(str(course).split()).split()
    if len(parts) < 2:
        return None
    subj = parts[0].upper()
    num = parts[-1].upper()
    if not re.match(r"^\d+[A-Z]?L?$", num):
        return None
    if num.endswith("L") and len(num) > 1:
        base = num[:-1]
        return f"{subj} {base}" if base else None
    return f"{subj} {num}L"


def _catalog_number_ends_with_l(course: str) -> bool:
    parts = str(course).split()
    if len(parts) < 2:
        return False
    num = parts[-1].upper()
    return num.endswith("L") and len(num) > 1


def _courses_to_remove_with_lab_partner(course: str) -> list[str]:
    s = " ".join(str(course).split()).strip()
    if not s:
        return []
    cand = [s]
    p = _lab_partner_course_code(s)
    if p:
        cand.append(p)
    seen: set[str] = set()
    uniq: list[str] = []
    for c in cand:
        ks = _norm_course_keys(c)
        if not ks:
            continue
        tag = min(ks)
        if tag in seen:
            continue
        seen.add(tag)
        uniq.append(c.strip())
    return uniq


def _lookup_course_snapshot(snapshots: dict[str, dict] | None, course: str) -> dict | None:
    if not snapshots:
        return None
    tgt = _norm_course_keys(course)
    for k, v in snapshots.items():
        if tgt & _norm_course_keys(k):
            return v
    return None


def _snapshot_from_enriched(enriched: list, course: str) -> dict | None:
    tgt = _norm_course_keys(course)
    for it in enriched or []:
        if not isinstance(it, dict):
            continue
        c = (it.get("course") or "").strip()
        if c and tgt & _norm_course_keys(c):
            return {"category": it.get("category"), "units": it.get("units")}
    return None


def _lecture_before_lab_pair(a: dict, b: dict) -> tuple[dict, dict]:
    ca = (a.get("course") or "").strip()
    cb = (b.get("course") or "").strip()
    a_lab = _catalog_number_ends_with_l(ca)
    b_lab = _catalog_number_ends_with_l(cb)
    if a_lab and not b_lab:
        return (b, a)
    if b_lab and not a_lab:
        return (a, b)
    return (a, b)


def _group_replacement_candidates_by_lab_pair(cands: list[dict]) -> list[list[dict]]:
    if not cands:
        return []
    by_norm: dict[str, dict] = {}
    for c in cands:
        course = (c.get("course") or "").strip()
        if not course:
            continue
        ks = _norm_course_keys(course)
        if not ks:
            continue
        by_norm[min(ks)] = c
    consumed: set[str] = set()
    groups: list[list[dict]] = []
    for c in cands:
        course_str = (c.get("course") or "").strip()
        ks = _norm_course_keys(course_str)
        if not ks or min(ks) in consumed:
            continue
        partner_code = _lab_partner_course_code(course_str)
        partner_obj: dict | None = None
        if partner_code:
            for pk in _norm_course_keys(partner_code):
                if pk in by_norm:
                    partner_obj = by_norm[pk]
                    break
        if partner_obj is not None:
            first, second = _lecture_before_lab_pair(c, partner_obj)
            for d in (first, second):
                for dk in _norm_course_keys((d.get("course") or "").strip()):
                    consumed.add(dk)
            groups.append([first, second])
        else:
            consumed.update(ks)
            groups.append([c])
    return groups


def _apply_calendar_plan_followup(
    user_id: int,
    gaps: list,
    pref: str,
    *,
    verify_ctx: dict | None = None,
    removed_snapshots: dict[str, dict] | None = None,
) -> None:
    """Re-run planning after a calendar-card removal request; refreshes enrichment cache."""
    prev = st.session_state.get("planning_result")
    if not isinstance(prev, dict):
        st.warning("No current plan to edit.")
        return
    removed_all: list[str] = []
    if verify_ctx:
        ra = verify_ctx.get("removed_all")
        if isinstance(ra, list) and ra:
            removed_all = [str(x).strip() for x in ra if str(x).strip()]
        else:
            one = str((verify_ctx.get("removed") or "")).strip()
            if one:
                removed_all = [one]
    augmented = list(gaps or [])
    for code in removed_all:
        snap = _lookup_course_snapshot(removed_snapshots, code)
        augmented = _augment_gaps_with_removed(augmented, code, snap)
    with st.spinner("Updating schedule…"):
        try:
            new_plan = plan_for_user(
                user_id,
                augmented,
                pref,
                previous_plan=prev,
            )
        except Exception as e:
            st.error(f"Replace failed: {e}")
            return
    removed_for_verify: str | list[str] = removed_all if removed_all else ""
    if verify_ctx:
        rows = verify_calendar_replacements(
            old_plan=prev,
            new_plan=new_plan,
            gaps=augmented,
            removed_course=removed_for_verify,
            vacated_col_i=verify_ctx.get("col_i"),
            vacated_parsed=verify_ctx.get("parsed"),
            base_schedule_map=verify_ctx.get("base_map") or {},
        )
        st.session_state["calendar_replace_verify"] = rows
    else:
        st.session_state.pop("calendar_replace_verify", None)

    base_map = (verify_ctx or {}).get("base_map") or {}
    col_i = verify_ctx.get("col_i") if verify_ctx else None
    parsed_slot = verify_ctx.get("parsed") if verify_ctx else None
    flat_cands = gap_courses_matching_slot(augmented, base_map, col_i, parsed_slot)
    st.session_state["calendar_replacement_suggestions"] = _group_replacement_candidates_by_lab_pair(
        flat_cands
    )
    st.session_state["calendar_replacement_slot_label"] = _slot_label_from_verify_ctx(verify_ctx)

    st.session_state["planning_result"] = new_plan
    st.session_state["last_planning_message"] = pref
    st.session_state.pop("enriched_courses", None)
    st.session_state.pop("_recommended_enrichment_fp", None)
    st.rerun()


if run and xlsx_file is None:
    st.warning("Please upload an xlsx file first.")

data = None
if run and xlsx_file is not None:
    data = parse_academic_progress_xlsx(xlsx_file.getvalue())
    parsed_rows = list(data.get("detail_rows", []))
    st.session_state["parsed_rows"] = parsed_rows
    st.session_state["parsed_course_codes"] = list(data.get("course_codes", []))
    st.session_state["missing_details"] = [
        {
            "course": row["course_code"],
            "category": row["requirement"],
            "units": row["units"],
        }
        for row in parsed_rows
        if row["status"] == "Not Satisfied"
    ]
    st.session_state["transcript_progress_snapshot"] = {
        "requirement_status_counts": dict(data.get("requirement_status_counts") or {}),
        "not_satisfied": list(data.get("not_satisfied") or []),
        "course_codes": list(data.get("course_codes") or []),
    }
    try:
        memory_agent.save_last_transcript_snapshot(
            USER_ID,
            {
                "missing_details": st.session_state["missing_details"],
                "parsed_rows": st.session_state["parsed_rows"],
                "requirement_status_counts": st.session_state["transcript_progress_snapshot"][
                    "requirement_status_counts"
                ],
                "not_satisfied": st.session_state["transcript_progress_snapshot"]["not_satisfied"],
                "course_codes": st.session_state["parsed_course_codes"],
            },
        )
    except Exception:
        pass
    st.session_state.pop("pdf_gap_result", None)
elif "parsed_rows" in st.session_state:
    data = _session_transcript_data_for_display()

if data:
    st.subheader("Step 1 · Progress overview")
    if not (run and xlsx_file is not None):
        st.info(
            "Using your previously uploaded transcript. Upload a new file below to update."
        )
    counts = data.get("requirement_status_counts", {})

    st.markdown("##### One merged status per DegreeWorks requirement block")
    c1, c2, c3 = st.columns(3)
    c1.metric("Satisfied blocks", counts.get("Satisfied", 0))
    c2.metric("In progress blocks", counts.get("In Progress", 0))
    c3.metric("Not satisfied blocks", counts.get("Not Satisfied", 0))

    st.subheader("Requirements still not satisfied")
    ns = data.get("not_satisfied", [])
    if ns:
        st.dataframe(_not_satisfied_table(ns), use_container_width=True, hide_index=True)
    else:
        st.success("No requirement blocks with status Not Satisfied (per Excel).")

    st.subheader("Parsed course codes from the sheet (unique)")
    codes = data.get("course_codes", [])
    st.write(", ".join(codes) if codes else "(No registration rows with a parseable course code)")

    st.subheader("All detail rows (aligned with Excel)")
    detail = list(data.get("detail_rows", []))
    if hide_empty:
        detail = [r for r in detail if r.get("registration")]
    if detail:
        st.dataframe(_detail_table_rows(detail), use_container_width=True, hide_index=True)
    else:
        st.info("No detail rows — headers may not match or the workbook format changed.")

st.divider()

pdf_gap = st.session_state.get("pdf_gap_result")
if isinstance(pdf_gap, dict):
    st.subheader("Curriculum PDF gap summary (last run)")
    comp = pdf_gap.get("completed") or []
    miss = pdf_gap.get("missing") or []
    if comp:
        st.markdown("**Satisfied (per PDF + your completed list)**")
        st.write(", ".join(str(x) for x in comp) if comp else "—")
    if miss:
        st.markdown("**Still required (per PDF)**")
        st.write(", ".join(str(x) for x in miss))
    md_show = pdf_gap.get("missing_details") or []
    if md_show:
        st.dataframe(
            [
                {
                    "Course": (i or {}).get("course"),
                    "Category": (i or {}).get("category"),
                    "Units": (i or {}).get("units"),
                }
                for i in md_show
                if isinstance(i, dict)
            ],
            use_container_width=True,
            hide_index=True,
        )

st.divider()
st.subheader("Step 2: Generate a recommended schedule (missing_details + your preferences)")

missing_details = st.session_state.get("missing_details")
if isinstance(missing_details, list) and missing_details:
    if "planning_user_preference" not in st.session_state:
        st.session_state["planning_user_preference"] = ""

    pending_append = st.session_state.pop("_pending_transcription_append", None)
    if pending_append:
        prev = (st.session_state.get("planning_user_preference") or "").strip()
        st.session_state["planning_user_preference"] = (
            f"{prev} {pending_append}".strip() if prev else pending_append
        )

    user_preference = st.text_area(
        "Describe your preferences (target units, time of day, which areas to prioritize…)",
        placeholder="e.g. at most 12 units, finish core first, no classes before 9am",
        key="planning_user_preference",
    )

    st.caption(
        "**Voice (optional):** record a short clip, then transcribe—it appends to the box above "
        "(English; uses Google speech recognition over the **network**)."
    )
    rec = st.audio_input("Record preference (optional)", key="planning_pref_audio")
    if st.button("Transcribe recording → preferences", key="planning_pref_transcribe"):
        if rec is None:
            st.warning("Record audio first, then click transcribe.")
        else:
            raw = rec.getvalue()
            text, err = transcribe_wav_bytes(raw)
            if err:
                st.warning(err)
            elif text:
                st.session_state["_pending_transcription_append"] = text
                st.success("Transcription added to your preferences.")
                st.rerun()

    if st.button("Generate recommended schedule"):
        if not (user_preference or "").strip():
            st.warning("Please enter your preferences first.")
        else:
            with st.spinner("Generating recommended schedule…"):
                try:
                    previous_plan = st.session_state.get("planning_result")
                    planning_result = plan_for_user(
                        USER_ID,
                        missing_details,
                        user_preference,
                        previous_plan=previous_plan if isinstance(previous_plan, dict) else None,
                    )
                    st.session_state["planning_result"] = planning_result
                    st.session_state.pop("calendar_replacement_suggestions", None)
                    st.session_state.pop("calendar_replacement_slot_label", None)
                    # Remember the preference text so the chat bubble can
                    # render the user side of the conversation alongside
                    # the AI's assistant_reply.
                    st.session_state["last_planning_message"] = (user_preference or "").strip()
                except Exception as e:
                    st.error(f"Generation failed: {e}")

    planning_result = st.session_state.get("planning_result")
    if isinstance(planning_result, dict):
        cal_verify = st.session_state.get("calendar_replace_verify")
        if cal_verify:
            with st.expander("Replacement vs Find Course Sections (last calendar action)", expanded=True):
                st.caption(
                    "New course codes compared to the Find Course Sections workbook; when the "
                    "removed card had a known pattern, **Slot fits vacated window** requires same "
                    "weekday column + overlapping clock times on that day."
                )
                st.dataframe(cal_verify, use_container_width=True, hide_index=True)

        plan_warning_msgs = _planning_warning_messages(planning_result)
        raw_recs = planning_result.get("recommended") or []
        base_schedule_map = _load_schedule_map_base_from_xlsx()
        if base_schedule_map:
            recs = [r for r in raw_recs if _recommended_item_in_find_course_xlsx(r, base_schedule_map)]
        else:
            recs = list(raw_recs)

        if recs:
            fp = _recommended_fingerprint(recs)
            if st.session_state.get("_recommended_enrichment_fp") != fp:
                with st.spinner("Fetching ratings from RateMyProfessor…"):
                    st.session_state["enriched_courses"] = run_professor_agent(recs)
                    st.session_state["_recommended_enrichment_fp"] = fp
        else:
            st.session_state.pop("enriched_courses", None)
            st.session_state.pop("_recommended_enrichment_fp", None)

        enriched = st.session_state.get("enriched_courses")

        assistant_reply = (planning_result.get("assistant_reply") or "").strip()
        last_user_msg = st.session_state.get("last_planning_message") or ""
        if assistant_reply or last_user_msg:
            if last_user_msg:
                with st.chat_message("user"):
                    st.write(last_user_msg)
            with st.chat_message("assistant"):
                if assistant_reply:
                    st.write(assistant_reply)
                else:
                    # Fallback when the model only filled `advice` (older
                    # mocks / response paths). Still better than silence.
                    fallback = (planning_result.get("advice") or "").strip()
                    st.write(fallback or "Updated the plan based on your preferences.")

        if plan_warning_msgs:
            st.markdown("##### Plan notices")
            for m in plan_warning_msgs:
                st.warning(m)

        if recs:
            st.subheader("Professor ratings (RateMyProfessor)")
            st.caption(
                "After a successful **Generate recommended schedule**, ratings load automatically. "
                "Each **left** card lists **all instructors returned for that course**, sorted by **rating high → low** "
                "(rating, difficulty, would-take-again). The row aligned to **Top pick** is the schedule/heuristic choice when present."
            )

        left, right = st.columns(2)

        with left:
            st.markdown("#### Recommended courses")
            if not recs:
                st.info("(Model returned no recommended list)")
            elif not isinstance(enriched, list):
                st.info("(No professor enrichment loaded)")
            else:
                for item in enriched:
                    with st.container(border=True):
                        course = item.get("course", "(Unknown course)")
                        category = item.get("category", "(Unknown category)")
                        units = item.get("units", "(Unknown units)")
                        reason = item.get("reason", "")

                        st.markdown(f"**{course}**  ·  {category}")
                        st.metric("Units", units)
                        sched = item.get("scheduled_instructors")
                        if isinstance(sched, list) and sched:
                            st.caption("Find Course instructors: " + ", ".join(sched))
                        if reason:
                            st.info(reason)

                        profs = item.get("professors") or []
                        err_msg = item.get("error")
                        rmp_note = item.get("rmp_note")

                        if err_msg:
                            st.warning(err_msg)
                        if rmp_note:
                            st.caption(rmp_note)

                        if not profs:
                            if not err_msg:
                                st.warning("No rating data")
                            continue

                        best_name = (item.get("best_professor") or "").strip()
                        ranked = professors_sorted_by_rating(profs)
                        st.markdown("**Instructors (rating high → low)**")
                        table_rows = []
                        for i, p in enumerate(ranked, start=1):
                            nm = (p.get("name") or "").strip() or "—"
                            note = ""
                            if best_name and nm == best_name:
                                note = "Top pick"
                            stars, numeric = _rating_display(p.get("rating"))
                            rating_cell = f"{stars} {numeric}".strip() if stars else numeric
                            diff = p.get("difficulty")
                            diff_txt = (
                                f"{diff:.1f}"
                                if isinstance(diff, (int, float))
                                else (str(diff) if diff is not None else "—")
                            )
                            wta = p.get("would_take_again") or "N/A"
                            table_rows.append(
                                {
                                    "#": i,
                                    "Instructor": nm,
                                    "Rating": rating_cell,
                                    "Difficulty": diff_txt,
                                    "Would take again": wta,
                                    "Note": note,
                                }
                            )
                        st.dataframe(table_rows, use_container_width=True, hide_index=True)

        with right:
            st.markdown("#### Summary")
            _tu = planning_result.get("total_units")
            if raw_recs:
                su = 0.0
                any_u = False
                for r in raw_recs:
                    try:
                        su += float(r.get("units"))
                        any_u = True
                    except (TypeError, ValueError):
                        pass
                if any_u:
                    st.metric("Total units", int(su) if su == int(su) else su)
                else:
                    st.metric("Total units", _tu if _tu is not None else "—")
            else:
                st.metric("Total units", _tu if _tu is not None else "—")
            advice = (planning_result.get("advice") or "").strip()
            if advice:
                st.write(advice)
            else:
                st.info("(Model returned no advice)")
            if plan_warning_msgs:
                st.markdown("**Plan notices**")
                for m in plan_warning_msgs:
                    st.warning(m)

        if recs and isinstance(enriched, list) and enriched:
            st.session_state["course_schedule_map"] = load_course_schedule_map_from_xlsx()
            schedule_map = st.session_state["course_schedule_map"]

            st.divider()
            st.subheader("Step 3 · Schedule preview")
            if plan_warning_msgs:
                for m in plan_warning_msgs:
                    st.warning(m)
            st.caption(
                "Times come from the Find Course Sections workbook (Meeting Patterns) in this folder; "
                "if a course has no matching section, it is listed under **Time TBD**. "
                "**Remove & replace from gaps** calls the planner again with your Step 1 gap list, "
                "asking for a replacement that prefers the same weekday and time window when known."
            )

            day_map = {"M": 0, "T": 1, "W": 2, "Th": 3, "R": 3, "F": 4}
            day_headers = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
            buckets: list[list[tuple[dict, dict]]] = [[] for _ in range(5)]
            pending_cal: list[dict] = []

            for item in enriched:
                code = (item.get("course") or "").strip()
                raw = schedule_map.get(code) if code else None
                parsed = parse_schedule(raw) if raw else None
                if not parsed:
                    pending_cal.append(item)
                    continue
                day_indices: list[int] = []
                for d in parsed["days"]:
                    idx = day_map.get(d)
                    if idx is not None:
                        day_indices.append(idx)
                if not day_indices:
                    pending_cal.append(item)
                    continue
                primary_col = min(day_indices)
                buckets[primary_col].append((item, parsed))

            header_cols = st.columns(5)
            for i, hc in enumerate(header_cols):
                with hc:
                    st.markdown(f"**{day_headers[i]}**")

            body_cols = st.columns(5)
            for col_i, bc in enumerate(body_cols):
                with bc:
                    for bi, (item, parsed) in enumerate(buckets[col_i]):
                        with st.container(border=True):
                            st.markdown(f"**{item.get('course', '(Unknown course)')}**")
                            st.caption(f"{parsed['start']} – {parsed['end']}")
                            st.write(item.get("best_professor") or "—")
                            fk = hashlib.sha256(
                                f"{col_i}|{bi}|{item.get('course', '')}|{parsed.get('start')}|{parsed.get('end')}".encode()
                            ).hexdigest()[:24]
                            if st.button(
                                "Remove & replace from gaps",
                                key=f"cal_rr_{fk}",
                                help=(
                                    "Drop this course and ask the planner for a replacement from your "
                                    "Step 1 gap list, preferring the same weekday/time window."
                                ),
                            ):
                                to_rm = _courses_to_remove_with_lab_partner(
                                    str(item.get("course") or "")
                                )
                                snapshots = {}
                                for code in to_rm:
                                    sn = _snapshot_from_enriched(enriched, code)
                                    snapshots[code] = (
                                        sn if sn else {"category": "Requirement gap", "units": 0}
                                    )
                                pref = build_remove_and_replace_preference(
                                    to_rm,
                                    day_headers[col_i],
                                    parsed,
                                )
                                _apply_calendar_plan_followup(
                                    USER_ID,
                                    missing_details,
                                    pref,
                                    verify_ctx={
                                        "removed": to_rm[0],
                                        "removed_all": to_rm,
                                        "col_i": col_i,
                                        "parsed": parsed,
                                        "base_map": _load_schedule_map_base_from_xlsx(),
                                        "day_label": day_headers[col_i],
                                    },
                                    removed_snapshots=snapshots,
                                )

            if pending_cal:
                st.markdown("##### Time TBD")
                for pi, item in enumerate(pending_cal):
                    with st.container(border=True):
                        st.markdown(f"**{item.get('course', '(Unknown course)')}**")
                        st.caption("No Meeting Patterns or no match in schedule xlsx")
                        st.write(item.get("best_professor") or "—")
                        fk = hashlib.sha256(
                            f"pend|{pi}|{item.get('course', '')}".encode()
                        ).hexdigest()[:24]
                        if st.button(
                            "Remove & replace from gaps",
                            key=f"cal_rr_{fk}",
                            help=(
                                "Drop this course and ask for a gap-based replacement (time window unknown)."
                            ),
                        ):
                            to_rm = _courses_to_remove_with_lab_partner(
                                str(item.get("course") or "")
                            )
                            snapshots = {}
                            for code in to_rm:
                                sn = _snapshot_from_enriched(enriched, code)
                                snapshots[code] = (
                                    sn if sn else {"category": "Requirement gap", "units": 0}
                                )
                            pref = build_remove_and_replace_preference(to_rm, None, None)
                            _apply_calendar_plan_followup(
                                USER_ID,
                                missing_details,
                                pref,
                                verify_ctx={
                                    "removed": to_rm[0],
                                    "removed_all": to_rm,
                                    "col_i": None,
                                    "parsed": None,
                                    "base_map": _load_schedule_map_base_from_xlsx(),
                                    "day_label": None,
                                },
                                removed_snapshots=snapshots,
                            )

            if "calendar_replacement_suggestions" in st.session_state:
                cand_groups = st.session_state.get("calendar_replacement_suggestions") or []
                with st.expander(
                    "Replacement suggestions for vacated slot",
                    expanded=True,
                ):
                    slot_lbl = st.session_state.get("calendar_replacement_slot_label") or ""
                    if slot_lbl:
                        st.caption(f"Freed time slot: **{slot_lbl}**")
                    if not cand_groups:
                        st.warning("No available courses found for this time slot")
                    else:
                        for gi, group in enumerate(cand_groups):
                            lines = []
                            for cand in group:
                                ccode = (cand.get("course") or "—").strip()
                                lines.append(
                                    f"**{ccode}** · {cand.get('units', '—')} units · "
                                    f"*{(cand.get('category') or '—').strip()}*"
                                )
                            if len(group) == 2:
                                st.markdown(
                                    "Lecture + lab pair:\n\n" + "\n\n".join(lines)
                                )
                            else:
                                st.markdown("\n\n".join(lines))
                            gkey = "|".join(
                                (c.get("course") or "").strip() for c in group
                            )
                            add_key = (
                                "cal_add_"
                                + hashlib.sha256(f"{gkey}|{gi}".encode()).hexdigest()[:24]
                            )
                            if st.button("Add to schedule", key=add_key):
                                pr = st.session_state.get("planning_result")
                                all_in = isinstance(pr, dict) and all(
                                    _course_already_in_recommended(pr, (c.get("course") or "").strip())
                                    for c in group
                                )
                                if all_in:
                                    st.info("These courses are already in your schedule.")
                                else:
                                    _append_courses_group_to_planning_result(group)
                                    st.session_state.pop("enriched_courses", None)
                                    st.session_state.pop("_recommended_enrichment_fp", None)
                                    st.rerun()
else:
    st.info(
        "(No `missing_details` yet: run Step 1 requirement analysis and set "
        "`st.session_state['missing_details']`.)"
    )
