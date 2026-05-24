#!/usr/bin/env python3
"""One-time scraper: pull SCU instructor ratings from RateMyProfessors.

Usage (run from project/course_planner/):
    python3 scripts/scrape_rmp_ratings.py

Reads  : SCU_Find_Course_Sections.xlsx  (instructor names)
Writes : data/instructor_ratings.csv    (overwrites placeholder rows)

Strategy
--------
For each unique instructor name in the schedule xlsx, search RMP by last name
and fuzzy-match the full name (same logic as _lookup_scheduled_professor in
professor_agent.py).  Instructors not found on RMP are silently skipped —
they resolve to rating=None at runtime and the InstructorSelector falls back
to a stable choice.

Rate-limit
----------
Sleeps _SLEEP_BETWEEN_S seconds between each request to avoid throttling.
With ~400 unique names that get looked up (~half hit RMP), expect ~5-10 min.
"""

from __future__ import annotations

import csv
import re
import sys
import time
from datetime import date
from pathlib import Path

# Locate course_planner root: prefer CWD if it looks right, else fall back
# to the directory two levels above this script file.
_CWD = Path.cwd()
_SCRIPT_ROOT = Path(__file__).resolve().parent.parent

def _find_root() -> Path:
    """Return the course_planner/ root that actually contains the xlsx."""
    for candidate in (_CWD, _SCRIPT_ROOT):
        if (candidate / "SCU_Find_Course_Sections.xlsx").is_file():
            return candidate
    # Last resort: walk up from CWD looking for the xlsx
    for parent in _CWD.parents:
        xlsx = parent / "course_planner" / "SCU_Find_Course_Sections.xlsx"
        if xlsx.is_file():
            return xlsx.parent
    raise FileNotFoundError(
        "Cannot find SCU_Find_Course_Sections.xlsx. "
        "Run this script from project/course_planner/."
    )

_HERE = _find_root()
if str(_HERE) not in sys.path:
    sys.path.insert(0, str(_HERE))

from rmp_client import RMPClient

_XLSX = _HERE / "SCU_Find_Course_Sections.xlsx"
_OUT  = _HERE / "data" / "instructor_ratings.csv"
_SCU_RMP_SCHOOL_ID = 882
_SLEEP_BETWEEN_S   = 0.4   # seconds between RMP requests


# ── helpers (mirrors professor_agent.py) ─────────────────────────────────────

def _names_same_person(rmp_name: str, schedule_name: str) -> bool:
    a = " ".join(rmp_name.lower().split())
    b = " ".join(schedule_name.lower().split())
    if a == b:
        return True
    ap, bp = a.split(), b.split()
    if not ap or not bp:
        return False
    if ap[-1] != bp[-1]:          # last names must match
        return False
    return ap[0][0] == bp[0][0]   # same first initial


def _lookup(client: RMPClient, school_id: int, schedule_name: str):
    """Return an RMP professor object or None."""
    query = schedule_name.split()[-1]   # search by last name
    if len(query) < 2:
        return None
    try:
        r = client.list_professors_for_school(school_id, query=query, page_size=20)
    except Exception as exc:
        print(f"    ⚠ RMP error for '{schedule_name}': {exc}")
        return None
    for p in r.professors or []:
        if _names_same_person(p.name, schedule_name):
            return p
    return None


# ── read instructor names from xlsx ──────────────────────────────────────────

def _extract_names(xlsx_path: Path) -> list[str]:
    from openpyxl import load_workbook
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c).strip() if c else "" for c in rows[0]]
    inst_idx = next(
        (i for i, h in enumerate(headers) if "instructor" in h.lower()), None
    )
    wb.close()
    if inst_idx is None:
        raise RuntimeError("No instructor column found in xlsx")

    seen: set[str] = set()
    names: list[str] = []
    for row in rows[1:]:
        cell = row[inst_idx] if inst_idx < len(row) else None
        if not cell:
            continue
        for part in re.split(r"[\|\n]", str(cell)):
            part = part.strip()
            if (
                part
                and len(part) > 3
                and part.lower() not in ("staff", "tba", "tbd", "none", "nan")
                and part not in seen
            ):
                seen.add(part)
                names.append(part)
    return names


# ── CSV header / comment ─────────────────────────────────────────────────────

_HEADER_COMMENT = f"""\
# instructor_ratings.csv — instructor quality data for R5 (best-section picker)
#
# PROVENANCE:
#   Scraped from RateMyProfessors.com on {date.today().isoformat()} for
#   Santa Clara University (school id {_SCU_RMP_SCHOOL_ID}).
#   Source column: "rmp" = live RMP data.
#   Instructors absent from RMP are not listed here; they resolve to
#   rating=None at runtime (InstructorSelector falls back to a stable choice).
#
# Columns:
#   instructor_name          exact name as it appears in the schedule xlsx
#   rating                   0.0-5.0 overall quality (higher = better)
#   difficulty               0.0-5.0 (lower = easier); tie-breaker only
#   would_take_again_pct     0-100
#   source                   provenance tag
"""

_CSV_FIELDS = [
    "instructor_name",
    "rating",
    "difficulty",
    "would_take_again_pct",
    "source",
]


# ── main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    print(f"Reading instructors from {_XLSX.name} …")
    names = _extract_names(_XLSX)
    print(f"  {len(names)} unique instructor names found.\n")

    client = RMPClient()
    # Resolve the live school id (falls back to hardcoded 882)
    try:
        res = client.search_schools("Santa Clara University", page_size=5)
        school_id = int(res.schools[0].id) if res.schools else _SCU_RMP_SCHOOL_ID
    except Exception:
        school_id = _SCU_RMP_SCHOOL_ID
    print(f"RMP school id: {school_id}\n")

    rows: list[dict] = []
    found = 0
    not_found = 0

    for i, name in enumerate(names, 1):
        prefix = f"[{i:3d}/{len(names)}]"
        prof = _lookup(client, school_id, name)
        if prof is None:
            print(f"{prefix} ✗ not found   {name}")
            not_found += 1
        else:
            rating  = prof.overall_rating
            diff    = prof.level_of_difficulty
            wta_raw = getattr(prof, "percent_take_again", None)
            wta     = round(float(wta_raw)) if wta_raw is not None else None
            rows.append(
                {
                    "instructor_name":     name,
                    "rating":              f"{float(rating):.1f}" if rating is not None else "",
                    "difficulty":          f"{float(diff):.1f}"   if diff    is not None else "",
                    "would_take_again_pct": str(wta) if wta is not None else "",
                    "source":              "rmp",
                }
            )
            print(
                f"{prefix} ✓ {name:40s}  "
                f"rating={rating}  diff={diff}  wta={wta}%"
            )
            found += 1
        time.sleep(_SLEEP_BETWEEN_S)

    # Write CSV
    _OUT.parent.mkdir(parents=True, exist_ok=True)
    with _OUT.open("w", newline="", encoding="utf-8") as fh:
        fh.write(_HEADER_COMMENT)
        writer = csv.DictWriter(fh, fieldnames=_CSV_FIELDS)
        writer.writeheader()
        writer.writerows(rows)

    print(f"\n{'='*60}")
    print(f"Done.  Found: {found}  Not found: {not_found}")
    print(f"Wrote {len(rows)} rows → {_OUT}")


if __name__ == "__main__":
    main()
