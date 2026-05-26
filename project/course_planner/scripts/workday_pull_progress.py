#!/usr/bin/env python3
"""Pull *View My Academic Progress* from Workday and upload or save the export.

Run from ``project/course_planner/`` (see ``scripts/README.md``).

Flow
----
1. Open a **headed** Chromium with a persistent profile (``.workday_profile/``).
2. **You** complete SCU SSO + Duo in that window — this script never sees credentials.
3. Navigate to *View My Academic Progress*, click Workday's native **Export to Excel**.
4. Validate the download in-process; abort loudly if parsing yields no rows.
5. Either POST to the local API (``--user-id``) or write the ``.xlsx`` (``--save``).

Exit codes
----------
* ``0`` — success
* ``1`` — validation failed or API upload failed
* ``2`` — browser launch, login timeout, navigation, or export failed

Cron / automation
-----------------
Everything after login is automatable. The first run (or after session expiry) still
requires a human at the keyboard for SSO + Duo. Keep the API running for POST mode.

Environment
-----------
* ``SCU_WORKDAY_URL`` — override the default Academic Progress task URL.
* ``WORKDAY_LOGIN_MANUAL=1`` — after login, press Enter in the terminal (skip URL auto-detect).
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
from collections.abc import Callable
from pathlib import Path
from typing import Any

# Locate course_planner root (same pattern as scrape_rmp_ratings.py).
_CWD = Path.cwd()
_SCRIPT_ROOT = Path(__file__).resolve().parent.parent


def _find_root() -> Path:
    for candidate in (_CWD, _SCRIPT_ROOT):
        if (candidate / "utils").is_dir() and (candidate / "scripts").is_dir():
            return candidate
    for parent in _CWD.parents:
        cp = parent / "course_planner"
        if (cp / "utils").is_dir():
            return cp
    raise FileNotFoundError(
        "Cannot find course_planner root. Run from project/course_planner/."
    )


_HERE = _find_root()
if str(_HERE) not in sys.path:
    sys.path.insert(0, str(_HERE))

from playwright.sync_api import Page, TimeoutError as PWTimeout  # noqa: E402

from scripts._workday_browser import (  # noqa: E402
    export_controls_visible,
    export_document_modal_visible,
    export_to_excel,
    launch,
    wait_for_login,
)
from utils.academic_progress_helpers import enrich_missing_details  # noqa: E402
from utils.academic_progress_xlsx import parse_academic_progress_xlsx, sanitize_parsed_rows  # noqa: E402

# ── Config ───────────────────────────────────────────────────────────────────

TASK_URL = os.environ.get(
    "SCU_WORKDAY_URL",
    # SCU menu opens this task id (2024+); override via env if Workday changes it again.
    "https://www.myworkday.com/scu/d/task/2998$29782.htmld",
)
DEFAULT_PROFILE_DIR = _HERE / ".workday_profile"
DEFAULT_UPLOAD_URL = "http://localhost:8000/api/upload/transcript"

NAV_TIMEOUT_MS = 60 * 1000
RENDER_WAIT_MS = 4_000  # SPA paint — proven stable for SCU Workday

_ACADEMIC_PROGRESS_HEADINGS = (
    "view my academic progress",
    "academic progress",
    "my academic progress",
)
_SEARCH_QUERY = "View My Academic Progress"


class ProgressValidationError(Exception):
    """Parsed Workday export is empty — layout likely changed."""


def _print_xlsx_preview(xlsx_bytes: bytes, *, max_rows: int = 8) -> None:
    """Print the first rows of each sheet to stderr (diagnose layout changes)."""
    from io import BytesIO

    from openpyxl import load_workbook

    try:
        wb = load_workbook(BytesIO(xlsx_bytes), read_only=False, data_only=True)
    except Exception as exc:  # noqa: BLE001
        print(f"Could not open xlsx for preview: {exc}", file=sys.stderr)
        return
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            print(f"  sheet {name!r}:", file=sys.stderr)
            for i, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True)):
                print(f"    [{i}] {row}", file=sys.stderr)
    finally:
        wb.close()


# ── Navigation (recovered from utils/workday_scraper.py @ c095321^) ───────────


def _wait_for_workday_content(
    page: Page, ready: Callable[[], bool] | None = None
) -> None:
    """Let the Workday SPA paint after a navigation.

    With ``ready`` given, return as soon as it reports the next view is up
    (polled), capped at ``RENDER_WAIT_MS``; otherwise fall back to the fixed
    pause that older runs relied on. The cap is counted as summed poll steps
    (not wall clock) so it still terminates promptly when ``wait_for_timeout``
    is a test no-op.
    """
    try:
        page.wait_for_load_state("domcontentloaded", timeout=30_000)
    except PWTimeout:
        pass
    if ready is None:
        page.wait_for_timeout(RENDER_WAIT_MS)
        return
    waited = 0
    step = 150
    while waited < RENDER_WAIT_MS:
        try:
            if ready():
                return
        except Exception:  # noqa: BLE001
            pass
        page.wait_for_timeout(step)
        waited += step


def _any_visible(page: Page, *texts: str) -> Callable[[], bool]:
    """Build a lightweight readiness check: any of ``texts`` currently visible.

    Uses ``is_visible()`` with no timeout (an instant DOM check), so it is cheap
    to poll and never adds its own waiting.
    """

    def _check() -> bool:
        for t in texts:
            try:
                if page.get_by_text(t, exact=False).first.is_visible():
                    return True
            except Exception:  # noqa: BLE001
                continue
        return False

    return _check


def _student_selection_modal_visible(page: Page) -> bool:
    """Workday 'Student' confirm dialog before the report — scoped to a dialog, not page body."""
    try:
        dialog = page.get_by_role("dialog").filter(
            has=page.get_by_text("Student", exact=True)
        )
        if not dialog.first.is_visible(timeout=1_000):
            return False
        return dialog.get_by_role("button", name="OK").first.is_visible(timeout=800)
    except Exception:  # noqa: BLE001
        pass
    try:
        popup = page.locator('[data-automation-id="wd-popup"]').filter(
            has_text="Student"
        )
        if popup.first.is_visible(timeout=800):
            return popup.get_by_role("button", name="OK").first.is_visible(timeout=800)
    except Exception:  # noqa: BLE001
        pass
    return False


def _dismiss_student_selection_modal(page: Page) -> bool:
    """Click OK on the student confirmation dialog if present."""
    if not _student_selection_modal_visible(page):
        return False
    print("Confirming student on Workday prompt (clicking OK)…", flush=True)
    scopes = [page.get_by_role("dialog").filter(has=page.get_by_text("Student", exact=True))]
    try:
        scopes.append(page.locator('[data-automation-id="wd-popup"]').filter(has_text="Student"))
    except Exception:  # noqa: BLE001
        pass
    for scope in scopes:
        for target in (
            scope.get_by_role("button", name="OK").first,
            scope.locator('[data-automation-id="uic_primaryButton"]').first,
            scope.locator('button:has-text("OK")').first,
            scope.locator('[data-automation-id="wd-CommandButton"]:has-text("OK")').first,
        ):
            try:
                target.click(timeout=5_000, force=True)
                _wait_for_workday_content(page, lambda: _on_academic_progress_page(page))
                page.wait_for_timeout(1_000)
                if not _student_selection_modal_visible(page):
                    return True
            except Exception:  # noqa: BLE001
                continue
    return not _student_selection_modal_visible(page)


def _degree_audit_markers_visible(page: Page) -> bool:
    """Rows/labels that appear on the real progress report, not the Academics hub."""
    markers = (
        "Requirements Effective",
        "Requirements Not Satisfied",
        "Unused Registrations",
        "Last Evaluated",
    )
    for text in markers:
        try:
            if page.get_by_text(text, exact=False).first.is_visible(timeout=600):
                return True
        except Exception:  # noqa: BLE001
            continue
    return False


def _on_academic_progress_page(page: Page) -> bool:
    """True when the degree-audit report is open — not Academics hub or Student OK modal."""
    if _student_selection_modal_visible(page):
        return False
    if export_controls_visible(page) or export_document_modal_visible(page):
        return True
    if _degree_audit_markers_visible(page):
        return True
    try:
        title = (page.title() or "").lower()
        if any(h in title for h in _ACADEMIC_PROGRESS_HEADINGS):
            return True
        heading = page.evaluate(
            "() => (document.querySelector('h1, [role=\"heading\"]') || {}).textContent || ''"
        )
        return any(h in str(heading).lower() for h in _ACADEMIC_PROGRESS_HEADINGS)
    except Exception:  # noqa: BLE001
        return False


def _finalize_report_page(page: Page) -> bool:
    """Dismiss student OK prompt if needed; return whether the report view is ready."""
    _dismiss_student_selection_modal(page)
    return _on_academic_progress_page(page)


def _click_labels(
    page: Page,
    labels: tuple[str, ...],
    *,
    role: str | None = None,
    ready: Callable[[], bool] | None = None,
) -> bool:
    """Click the first visible control matching any of ``labels``.

    After a successful click, wait for the next view via ``ready`` (the element
    the *next* step needs) instead of a blanket fixed pause — the win that turns
    the ~4s-per-hop menu walk into "continue the moment it's painted".
    """
    for label in labels:
        candidates: list[Any] = []
        if role:
            try:
                candidates.append(page.get_by_role(role, name=label).first)
            except Exception:  # noqa: BLE001
                pass
        candidates.extend(
            [
                page.get_by_text(label, exact=True).first,
                page.locator(f'a:has-text("{label}")').first,
                page.locator(f'button:has-text("{label}")').first,
                page.locator(f'[aria-label*="{label}" i]').first,
            ]
        )
        for target in candidates:
            try:
                target.wait_for(state="visible", timeout=3_000)
                target.click(timeout=5_000)
                page.wait_for_timeout(400)
                _wait_for_workday_content(page, ready)
                return True
            except Exception:  # noqa: BLE001
                continue
    return False


def _click_academic_progress_link(page: Page) -> bool:
    """Click *View My Academic Progress* when its menu group is already open."""

    def _report_or_modal() -> bool:
        return _on_academic_progress_page(page) or _student_selection_modal_visible(page)

    clicked = _click_labels(
        page, ("View My Academic Progress",), role="link", ready=_report_or_modal
    ) or _click_labels(page, ("View My Academic Progress",), ready=_report_or_modal)
    if not clicked:
        return False
    return _finalize_report_page(page)


def _try_home_academics_app(page: Page) -> bool:
    """SCU home: Academics → View More → Academic Advising → View My Academic Progress (proven path)."""
    print(
        "Trying home: Academics → View More → Academic Advising → "
        "View My Academic Progress…",
        flush=True,
    )
    _wait_for_workday_content(page, _any_visible(page, "Academics", "SCU Academics"))

    after_academics = _any_visible(page, "Academic Advising", "View More", "View All Apps")
    if not _click_labels(
        page, ("Academics", "SCU Academics"), role="link", ready=after_academics
    ):
        if not _click_labels(page, ("Academics",), ready=after_academics):
            return False

    _click_labels(
        page,
        ("View More", "View All Apps", "View All Apps >", "View All"),
        ready=_any_visible(page, "Academic Advising"),
    )
    advising_ready = _any_visible(page, "View My Academic Progress")
    if not _click_labels(page, ("Academic Advising",), ready=advising_ready):
        _click_labels(page, ("Academic Advising",), role="button", ready=advising_ready)

    return _click_academic_progress_link(page)


def _try_sidebar_menu(page: Page) -> bool:
    """Legacy/alternate layout: Academic Advising group already visible in a side nav."""
    print("Trying sidebar: Academic Advising → View My Academic Progress…", flush=True)
    _wait_for_workday_content(
        page, _any_visible(page, "Academic Advising", "View My Academic Progress")
    )
    _click_labels(
        page, ("Academic Advising",), ready=_any_visible(page, "View My Academic Progress")
    )
    return _click_academic_progress_link(page)


def _try_search_for_report(page: Page) -> bool:
    search_box_selectors = [
        '[data-automation-id="globalSearchBox"]',
        '[data-automation-id="searchBox"]',
        '[data-automation-id="GLOBAL_SEARCH_BOX"]',
        'input[aria-label*="Search" i]',
        'input[placeholder*="Search" i]',
    ]
    box = None
    for sel in search_box_selectors:
        try:
            page.wait_for_selector(sel, timeout=4_000, state="visible")
            box = page.locator(sel).first
            box.click()
            break
        except Exception:  # noqa: BLE001
            box = None
            continue
    if box is None:
        return False

    def _report_ready() -> bool:
        return _on_academic_progress_page(page) or _student_selection_modal_visible(page)

    try:
        box.fill(_SEARCH_QUERY)
        page.wait_for_timeout(1_500)
        suggestion_selectors = [
            f'text="{_SEARCH_QUERY}"',
            'text="View My Academic Progress"',
            'a:has-text("View My Academic Progress")',
            '[data-automation-id="searchResults"] a',
        ]
        for sel in suggestion_selectors:
            try:
                page.locator(sel).first.click(timeout=3_000)
                _wait_for_workday_content(page, _report_ready)
                if _on_academic_progress_page(page):
                    return True
            except Exception:  # noqa: BLE001
                continue
        _wait_for_workday_content(page, _report_ready)
        return _on_academic_progress_page(page)
    except Exception:  # noqa: BLE001
        return False


def _ensure_on_task(page: Page, task_url: str = TASK_URL) -> None:
    """Navigate to View My Academic Progress (Academics menu first — matches successful runs)."""
    print("Navigating to View My Academic Progress…", flush=True)

    def _home_or_report() -> bool:
        if _on_academic_progress_page(page) or _student_selection_modal_visible(page):
            return True
        return _any_visible(page, "Academics", "SCU Academics")()

    _wait_for_workday_content(page, _home_or_report)
    if _finalize_report_page(page):
        print("Already on the Academic Progress report.", flush=True)
        return

    if _try_home_academics_app(page):
        print("Reached Academic Progress via Academics app menu.", flush=True)
        return

    if _try_sidebar_menu(page):
        print("Reached Academic Progress via sidebar menu.", flush=True)
        return

    print("Opening task URL…", flush=True)

    def _report_ready() -> bool:
        return _on_academic_progress_page(page) or _student_selection_modal_visible(page)

    try:
        page.goto(task_url, timeout=NAV_TIMEOUT_MS, wait_until="domcontentloaded")
        _wait_for_workday_content(page, _report_ready)
    except Exception:  # noqa: BLE001
        pass
    if _finalize_report_page(page):
        print("Reached Academic Progress via task URL.", flush=True)
        return

    if _try_home_academics_app(page):
        print("Reached Academic Progress via Academics app menu (after task URL).", flush=True)
        return

    if _try_sidebar_menu(page):
        print("Reached Academic Progress via sidebar menu (after task URL).", flush=True)
        return

    print("Task URL did not land on the report — trying Workday search…", flush=True)
    if _try_search_for_report(page):
        print("Reached Academic Progress via search.", flush=True)
        return

    actual = ""
    try:
        actual = (page.title() or "").strip()
    except Exception:  # noqa: BLE001
        pass
    raise RuntimeError(
        "Could not reach 'View My Academic Progress' in Workday"
        + (f" — currently on {actual!r}." if actual else ".")
        + "\nThe SCU_WORKDAY_URL task ID may be stale and the search fallback failed. "
        "Workaround: export manually in Workday and upload via the paperclip in the app."
    )


def navigate_to_academic_progress(page: Page) -> None:
    """``navigate`` callback for ``export_to_excel``."""
    _ensure_on_task(page)
    if _student_selection_modal_visible(page):
        _dismiss_student_selection_modal(page)
    if not _on_academic_progress_page(page):
        raise RuntimeError(
            "Opened View My Academic Progress but the report did not load after OK. "
            "Click OK on the Student dialog manually, then re-run."
        )


# ── Validation & upload ───────────────────────────────────────────────────────


def _prompt_return_to_planner(context: Any) -> None:
    """Brief pause + tab title so the user switches back to the Course Planner browser tab."""
    hint = "Done — switch back to your Course Planner tab (localhost:5173)"
    print(f"{hint}\n", flush=True)
    try:
        for p in context.pages:
            try:
                p.evaluate(f"document.title = {hint!r};")
            except Exception:  # noqa: BLE001
                pass
    except Exception:  # noqa: BLE001
        pass
    try:
        time.sleep(0.6)
    except Exception:  # noqa: BLE001
        pass


def persist_progress_for_user(user_id: str, parsed: dict[str, Any]) -> dict[str, Any]:
    """Write parsed progress to per-user memory (same as ``POST /api/upload/transcript``)."""
    from agents.memory_agent import write as memory_write

    parsed_rows = sanitize_parsed_rows(parsed.get("detail_rows") or [])
    missing_details = enrich_missing_details(parsed.get("not_satisfied") or [], parsed_rows)
    uid = user_id.strip()
    if uid:
        try:
            memory_write(uid, "academic_progress", json.dumps(missing_details))
        except Exception:  # noqa: BLE001
            pass
        try:
            memory_write(uid, "parsed_rows", json.dumps(parsed_rows))
        except Exception:  # noqa: BLE001
            pass
    return {"missing_details": missing_details, "parsed_rows": parsed_rows}


def pull_academic_progress(
    user_id: str,
    *,
    profile_dir: Path | None = None,
    progress_cb: Callable[[str], None] | None = None,
    manual_login: bool = False,
) -> dict[str, Any]:
    """Headed Workday pull: human SSO in browser, then export + parse + memory persist.

    ``progress_cb`` receives coarse status codes (``browser_open``, ``logged_in``, …)
    for UI polling. Raises ``ProgressValidationError``, ``TimeoutError``, or
    ``RuntimeError`` on failure.
    """

    def _cb(status: str) -> None:
        if progress_cb is not None:
            progress_cb(status)

    _cb("pending")
    context = None
    export_succeeded = False
    try:
        context, page = launch(profile_dir or DEFAULT_PROFILE_DIR)
        _cb("browser_open")
        if manual_login or os.environ.get("WORKDAY_LOGIN_MANUAL") == "1":
            print(
                "\nManual login mode: complete SSO + Duo in the browser, "
                "open the Workday home page, then press Enter here.\n",
                flush=True,
            )
            input()
            from scripts._workday_browser import _active_workday_page

            page = _active_workday_page(page) or page
            try:
                page.bring_to_front()
            except Exception:  # noqa: BLE001
                pass
        else:
            page = wait_for_login(page, progress_cb=_cb)
        if progress_cb is None:
            _cb("logged_in")

        def _navigate(page: Page) -> None:
            _cb("navigating")
            navigate_to_academic_progress(page)
            _cb("report_open")

        _cb("downloading")

        def _export_poll() -> None:
            _cb("exporting")

        xlsx_bytes = export_to_excel(page, _navigate, on_poll=_export_poll)
        export_succeeded = True
    finally:
        if context is not None:
            if export_succeeded:
                _prompt_return_to_planner(context)
            try:
                context.close()
            except Exception:  # noqa: BLE001
                pass

    _cb("parsing")
    parsed = validate_progress_export(xlsx_bytes)
    return persist_progress_for_user(user_id, parsed)


def validate_progress_export(xlsx_bytes: bytes) -> dict[str, Any]:
    """Parse export bytes; raise if both ``detail_rows`` and ``not_satisfied`` are empty."""
    data = parse_academic_progress_xlsx(xlsx_bytes)
    detail_rows = data.get("detail_rows") or []
    not_satisfied = data.get("not_satisfied") or []
    if not detail_rows and not not_satisfied:
        raise ProgressValidationError(
            "Workday export parsed to ZERO rows (detail_rows and not_satisfied both empty).\n"
            "The Academic Progress Excel layout may have changed — NOT writing or uploading "
            "this file. Re-export manually or update the parser."
        )
    return data


def _post_transcript(
    xlsx_bytes: bytes,
    *,
    user_id: str,
    upload_url: str = DEFAULT_UPLOAD_URL,
) -> dict[str, Any]:
    import requests

    files = {
        "file": (
            "academic_progress.xlsx",
            xlsx_bytes,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
    }
    data = {"user_id": user_id}
    resp = requests.post(upload_url, files=files, data=data, timeout=120)
    resp.raise_for_status()
    body = resp.json()
    if not isinstance(body, dict):
        raise RuntimeError(f"Unexpected API response type: {type(body)!r}")
    return body


def _print_upload_summary(body: dict[str, Any], parsed: dict[str, Any]) -> None:
    local_rows = len(parsed.get("detail_rows") or [])
    local_missing = len(parsed.get("not_satisfied") or [])
    missing = body.get("missing_details")
    if missing is None:
        detail_rows = parsed.get("detail_rows") or []
        not_satisfied = parsed.get("not_satisfied") or []
        missing = enrich_missing_details(not_satisfied, detail_rows)
    parsed_rows = body.get("parsed_rows") or []
    print(f"parsed locally: detail_rows={local_rows}, not_satisfied={local_missing}")
    print(f"API stored:       missing_details={len(missing)}, parsed_rows={len(parsed_rows)}")
    if local_rows and not parsed_rows:
        print(
            "\nWARNING: API returned empty parsed_rows but the xlsx parsed fine locally.\n"
            "Restart the API so it reloads course_planner (parser fix), then re-run:\n"
            "  cd project/api && uvicorn main:app --reload --port 8000 "
            "--reload-dir . --reload-dir ../course_planner\n",
            file=sys.stderr,
        )


# ── CLI ───────────────────────────────────────────────────────────────────────


def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Pull View My Academic Progress from Workday (headed browser, human login).",
    )
    p.add_argument(
        "--user-id",
        metavar="ID",
        help="User id for POST /api/upload/transcript (required unless --save).",
    )
    p.add_argument(
        "--save",
        metavar="PATH",
        type=Path,
        help="Write the .xlsx to PATH instead of uploading.",
    )
    p.add_argument(
        "--profile-dir",
        type=Path,
        default=DEFAULT_PROFILE_DIR,
        help=f"Persistent Chromium profile (default: {DEFAULT_PROFILE_DIR}).",
    )
    p.add_argument(
        "--upload-url",
        default=DEFAULT_UPLOAD_URL,
        help=f"Upload endpoint (default: {DEFAULT_UPLOAD_URL}).",
    )
    return p


def _pull_to_xlsx_bytes(
    profile_dir: Path,
    *,
    manual_login: bool = False,
) -> bytes:
    """Browser login + export only; returns raw xlsx bytes (CLI --save path)."""
    context = None
    try:
        context, page = launch(profile_dir)
        if manual_login or os.environ.get("WORKDAY_LOGIN_MANUAL") == "1":
            print(
                "\nManual login mode: complete SSO + Duo in the browser, "
                "open the Workday home page, then press Enter here.\n",
                flush=True,
            )
            input()
            from scripts._workday_browser import _active_workday_page

            page = _active_workday_page(page) or page
            try:
                page.bring_to_front()
            except Exception:  # noqa: BLE001
                pass
            print(f"Continuing with tab: {page.url}\n", flush=True)
        else:
            page = wait_for_login(page)
        print("Exporting to Excel (do not close the browser)…", flush=True)
        return export_to_excel(page, navigate_to_academic_progress)
    finally:
        if context is not None:
            try:
                context.close()
            except Exception:  # noqa: BLE001
                pass


def main(argv: list[str] | None = None) -> int:
    args = _build_parser().parse_args(argv)

    if args.save:
        do_upload = False
    elif args.user_id:
        do_upload = True
    else:
        _build_parser().error("one of --user-id or --save is required")

    exit_code = 0
    xlsx_bytes: bytes | None = None
    try:
        if args.save:
            xlsx_bytes = _pull_to_xlsx_bytes(
                args.profile_dir,
                manual_login=os.environ.get("WORKDAY_LOGIN_MANUAL") == "1",
            )
        else:
            result = pull_academic_progress(
                args.user_id.strip(),
                profile_dir=args.profile_dir,
                manual_login=os.environ.get("WORKDAY_LOGIN_MANUAL") == "1",
            )
            print(
                f"parsed locally: detail_rows={len(result.get('parsed_rows') or [])}, "
                f"not_satisfied={len(result.get('missing_details') or [])}"
            )
            print(
                f"API stored:       missing_details={len(result.get('missing_details') or [])}, "
                f"parsed_rows={len(result.get('parsed_rows') or [])}"
            )
            return 0
    except TimeoutError as exc:
        print(f"ERROR (login): {exc}", file=sys.stderr)
        exit_code = 2
    except RuntimeError as exc:
        print(f"ERROR (browser/export): {exc}", file=sys.stderr)
        exit_code = 2
    except ProgressValidationError as exc:
        print(f"\n{'=' * 60}\nVALIDATION FAILED\n{'=' * 60}", file=sys.stderr)
        print(str(exc), file=sys.stderr)
        return 1
    except Exception as exc:  # noqa: BLE001 — Playwright / launch failures
        print(f"ERROR (browser): {exc}", file=sys.stderr)
        exit_code = 2

    if exit_code != 0:
        return exit_code

    assert xlsx_bytes is not None
    try:
        parsed = validate_progress_export(xlsx_bytes)
    except ProgressValidationError as exc:
        debug_path = _HERE / ".workday_debug_last.xlsx"
        try:
            debug_path.write_bytes(xlsx_bytes)
            print(f"Debug export saved → {debug_path}", file=sys.stderr)
            _print_xlsx_preview(xlsx_bytes)
        except Exception:  # noqa: BLE001
            pass
        print(f"\n{'=' * 60}\nVALIDATION FAILED\n{'=' * 60}", file=sys.stderr)
        print(str(exc), file=sys.stderr)
        return 1
    except Exception as exc:  # noqa: BLE001
        print(f"ERROR (parse): {exc}", file=sys.stderr)
        return 1

    if do_upload:
        try:
            body = _post_transcript(
                xlsx_bytes,
                user_id=args.user_id.strip(),
                upload_url=args.upload_url,
            )
        except Exception as exc:  # noqa: BLE001
            print(f"ERROR (upload): {exc}", file=sys.stderr)
            return 1
        _print_upload_summary(body, parsed)
        api_rows = len(body.get("parsed_rows") or [])
        if not api_rows and (parsed.get("detail_rows") or []):
            return 1
        return 0

    out = args.save.expanduser().resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_bytes(xlsx_bytes)
    print(f"Saved {len(xlsx_bytes)} bytes → {out}")
    print(
        f"parsed: detail_rows={len(parsed.get('detail_rows') or [])}, "
        f"not_satisfied={len(parsed.get('not_satisfied') or [])}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
