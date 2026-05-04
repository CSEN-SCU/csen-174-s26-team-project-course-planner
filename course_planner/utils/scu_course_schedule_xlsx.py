"""
Parse SCU Find Course / Sections-style exports (.xlsx) to align recommended courses with **scheduled instructors**.

Default files (relative to the ``course_planner/`` package directory):
- ``SCU_Find_Course_Sections.xlsx``
- ``scu_find_course.xlsx``
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

_COURSE_PLANNER_DIR = Path(__file__).resolve().parents[1]
_DEFAULT_SCHEDULE_FILES = (
    _COURSE_PLANNER_DIR / "SCU_Find_Course_Sections.xlsx",
    _COURSE_PLANNER_DIR / "scu_find_course.xlsx",
)


def _find_schedule_path(explicit: Path | None) -> Path | None:
    if explicit is not None and explicit.is_file():
        return explicit
    for p in _DEFAULT_SCHEDULE_FILES:
        if p.is_file():
            return p
    return None


def _parse_section_subject_number(course_section: str | None) -> tuple[str, str] | None:
    """
    Parse (subject, catalog number) from the first segment of ``Course Section``.
    Example: ``CSEN 122-1 - Computer Architecture (-)`` -> ("CSEN", "122").
    """
    if not course_section or not isinstance(course_section, str):
        return None
    head = course_section.split(" - ")[0].strip().upper()
    m = re.match(r"^([A-Z]{2,8})\s+(\d+[A-Z]?)\s*-\s*\d+\s*$", head)
    if not m:
        return None
    return m.group(1), m.group(2)


def _normalize_planner_course_text(course_code: str) -> str:
    """Normalize ``153/L`` and extra whitespace for tokenization."""
    u = course_code.upper().replace("&", " ").replace(",", " ")
    u = re.sub(r"(\d+)\s*/\s*L\b", r"\1L", u)
    u = u.replace("/", " ")
    return " ".join(u.split())


def planned_section_keys(course_code: str) -> set[tuple[str, str]]:
    """
    Build the set of (SUBJECT, NUMBER) keys from a planner/recommended course string aligned with ``Course Section``.
    Cross-map COEN/CSEN and ECEN/ELEN to match common workbook spellings.
    """
    text = _normalize_planner_course_text(course_code)
    tokens = [t for t in text.split() if t]
    keys: set[tuple[str, str]] = set()
    i = 0
    while i < len(tokens):
        t = tokens[i]
        if (
            i + 1 < len(tokens)
            and re.fullmatch(r"[A-Z]{2,8}", t)
            and re.fullmatch(r"\d+[A-Z]?", tokens[i + 1])
        ):
            subj, num = t, tokens[i + 1]
            keys.add((subj, num))
            if subj == "COEN":
                keys.add(("CSEN", num))
            elif subj == "CSEN":
                keys.add(("COEN", num))
            elif subj == "ECEN":
                keys.add(("ELEN", num))
            elif subj == "ELEN":
                keys.add(("ECEN", num))
            i += 2
            continue
        i += 1
    return keys


def _split_instructor_aliases(cell: str | None) -> list[str]:
    if not cell or not isinstance(cell, str):
        return []
    return [p.strip() for p in re.split(r"\|", cell) if p.strip() and p.strip().lower() != "none"]


def load_schedule_section_index(path: Path | None = None) -> dict[tuple[str, str], list[str]]:
    """
    Read xlsx and build (subject abbr, catalog number) -> instructor name list (unique, order preserved).
    Uses columns ``Course Section`` and ``All Instructors`` only.
    """
    p = _find_schedule_path(path)
    if p is None:
        return {}

    index: dict[tuple[str, str], list[str]] = {}
    wb = load_workbook(p, read_only=True, data_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if not header:
            return {}
        h = [str(c).strip() if c is not None else "" for c in header]
        try:
            idx_sec = h.index("Course Section")
            idx_inst = h.index("All Instructors")
        except ValueError:
            return {}

        for row in it:
            if not row or idx_sec >= len(row):
                continue
            key = _parse_section_subject_number(row[idx_sec])  # type: ignore[arg-type]
            if not key:
                continue
            names = _split_instructor_aliases(row[idx_inst] if idx_inst < len(row) else None)  # type: ignore[arg-type]
            # Keep rows with no instructor (e.g. lecture) so lab instructors can merge into the base number
            bucket = index.setdefault(key, [])
            for n in names:
                if n not in bucket:
                    bucket.append(n)
    finally:
        wb.close()

    _merge_lab_instructors_into_base(index)
    _mirror_ecen_elen_keys(index)
    return index


def _merge_lab_instructors_into_base(
    index: dict[tuple[str, str], list[str]],
) -> None:
    """Merge instructors from ``SUBJ 153L`` into ``SUBJ 153`` (lecture rows often have None)."""
    for (subj, num) in list(index.keys()):
        s = str(num)
        if not s.endswith("L") or len(s) < 2:
            continue
        if not s[:-1].isdigit():
            continue
        base_num = s[:-1]
        base_key = (subj, base_num)
        base_list = index.setdefault(base_key, [])
        for n in index.get((subj, s), []):
            if n not in base_list:
                base_list.append(n)


def _mirror_ecen_elen_keys(index: dict[tuple[str, str], list[str]]) -> None:
    """When the workbook uses ECEN only, mirror the same list to ELEN (shared list object)."""
    for (subj, num) in list(index.keys()):
        if subj == "ECEN":
            alt = ("ELEN", num)
            if alt not in index:
                index[alt] = index[(subj, num)]
        elif subj == "ELEN":
            alt = ("ECEN", num)
            if alt not in index:
                index[alt] = index[(subj, num)]


def scheduled_instructors_for_course(
    course_code: str, index: dict[tuple[str, str], list[str]]
) -> list[str]:
    """All scheduled instructors for ``course_code`` across sections (merged)."""
    if not index:
        return []
    want = planned_section_keys(course_code)
    out: list[str] = []
    for k in want:
        for name in index.get(k, []):
            if name not in out:
                out.append(name)
    return out
