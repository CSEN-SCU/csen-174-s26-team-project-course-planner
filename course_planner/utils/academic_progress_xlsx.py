from __future__ import annotations

import re
from collections import defaultdict
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


def _merge_requirement_statuses(status_set: set[str]) -> str:
    if "Not Satisfied" in status_set:
        return "Not Satisfied"
    if "In Progress" in status_set:
        return "In Progress"
    return next(iter(sorted(status_set))) if status_set else ""

_CELL_NOISE_RE = (
    # 标题里可能出现的括号片段，解析课号时需去掉（课号在主段第一句）
    re.compile(r"\s*\([^)]*In Progress[^)]*\)\s*", re.IGNORECASE),
    re.compile(r"\s*\([^)]*Transfer Credit[^)]*\)\s*", re.IGNORECASE),
)


def registration_to_course_code(cell: str | None) -> str | None:
    """从「COEN 10 - Introduction …」这一类单元格里抽出课号，如 COEN 10 / CSEN 140L。"""
    if cell is None or not isinstance(cell, str):
        return None
    head = cell.split(" - ", 1)[0].strip()
    for rx in _CELL_NOISE_RE:
        head = rx.sub("", head).strip()
    parts = head.split()
    if len(parts) < 2:
        return None
    subj = parts[0].strip().upper()
    num = parts[1].strip().upper()
    if len(subj) < 2 or len(subj) > 8:
        return None
    if not re.fullmatch(r"[A-Z]{2,8}", subj):
        return None
    if not re.fullmatch(r"\d+[A-Z]*", num):
        return None
    return f"{subj} {num}"


def parse_academic_progress_xlsx(xlsx_bytes: bytes) -> dict[str, Any]:
    """解析 SCU「View My Academic Progress」导出表（单列 sheet）。

    返回结构中 ``detail_rows`` 适合直接喂给 ``st.dataframe``；
    ``not_satisfied`` 为仍为 Not Satisfied 的要求块简要信息；
    ``course_codes`` 为表中登记行解析出的全部课号（去重、排序）。
    """
    wb = load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    detail_rows: list[dict[str, Any]] = []
    not_satisfied: list[dict[str, Any]] = []
    all_codes: list[str] = []
    requirement_status_sets: defaultdict[str, set[str]] = defaultdict(set)
    requirement_status: dict[str, str] = {}

    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)

        header_found = False
        for row in it:
            if not header_found:
                if row and row[0] == "Requirement":
                    header_found = True
                continue
            if row is None or all(c is None or str(c).strip() == "" for c in row[:4]):
                continue

            requirement = row[0] if row[0] is not None else ""
            status = row[1] if row[1] is not None else ""
            remaining = row[2] if len(row) > 2 else None
            registration = row[3] if len(row) > 3 else None
            period = row[4] if len(row) > 4 else None
            units = row[5] if len(row) > 5 else None
            grade = row[6] if len(row) > 6 else None

            rq = str(requirement).strip()
            if not rq:
                continue
            st = str(status).strip()
            rm = remaining if remaining is None else (
                remaining if isinstance(remaining, (int, float)) else str(remaining).strip() or None
            )
            reg = registration if registration is None else (
                str(registration).strip() or None
            )

            if st:
                requirement_status_sets[rq].add(st)

            code = registration_to_course_code(reg)
            if code:
                all_codes.append(code)

            detail_rows.append(
                {
                    "requirement": rq,
                    "status": st,
                    "remaining": rm,
                    "registration": reg,
                    "course_code": code,
                    "academic_period": period,
                    "units": units,
                    "grade": grade,
                }
            )

        requirement_status = {
            rq: _merge_requirement_statuses(seen)
            for rq, seen in sorted(requirement_status_sets.items())
        }

        for rq, merged in requirement_status.items():
            if merged == "Not Satisfied":
                exemplar = next(
                    (
                        r
                        for r in detail_rows
                        if r["requirement"] == rq and r["status"] == "Not Satisfied"
                    ),
                    next((r for r in detail_rows if r["requirement"] == rq), {}),
                )
                not_satisfied.append(
                    {
                        "requirement": rq,
                        "remaining": exemplar.get("remaining"),
                        "status": "Not Satisfied",
                    }
                )

    finally:
        wb.close()

    course_codes = sorted(set(all_codes), key=lambda c: (c.split()[0], c.split()[1] if len(c.split()) > 1 else ""))

    stats: dict[str, int] = {}
    for rq, merged in requirement_status.items():
        stats[merged] = stats.get(merged, 0) + 1

    return {
        "detail_rows": detail_rows,
        "not_satisfied": not_satisfied,
        "course_codes": course_codes,
        "requirement_status": requirement_status,
        "requirement_status_counts": stats,
    }
